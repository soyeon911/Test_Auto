from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

_BLUE_DARK = "3F5F7F"      # 기존 1F497D보다 부드러운 남색
_BLUE_TITLE = "5B8DB8"     # 제목용 soft blue
_BLUE_LIGHT = "E8F1FA"     # 연한 파란 배경

_GREEN_DARK = "4F6F45"     # 진녹색이지만 덜 탁함
_GREEN_FILL = "F1F8F1"     # PASS 배경

_RED_FILL = "F9EEEE"       # FAIL 배경
_RED_FONT = "9E3A3A"       # 기존 C00000보다 덜 자극적인 빨강

_YELLOW_FILL = "FFF7D6"    # WARN 배경

_GRAY_HEADER = "EEEEEE"    # 헤더 회색 완화
_WHITE = "FFFFFF"
_STRIPE = "FBFBFB"         # 거의 흰색에 가까운 줄무늬
_METHOD_COLORS = {
    "GET": "EEF6EE",       # soft green
    "POST": "FDF2E6",      # soft orange
    "PUT": "FFF7D6",       # soft yellow
    "DELETE": "F9EEEE",    # soft red
    "PATCH": "F3ECF5",     # soft purple
}

_TYPE_KO = {
    "positive": "정상(Positive)",
    "wrong_type": "타입오류(WrongType)",
    "boundary": "경계값(Boundary)",
    "semantic_probe": "시맨틱(SemanticProbe)",
    "missing_required": "필수누락(MissingRequired)",
    "invalid_enum": "열거형오류(InvalidEnum)",
    "raw_image_relation": "이미지관계(RawImageRelation)",
}

_EXPECTED_RESULT_TYPE_KO = {
    "expected_pass": "기대: 성공",
    "expected_fail": "기대: 실패",
    "probe_only": "탐색용 (Probe Only)",
}

_AXIS_LABEL = {
    "schema": "schema (구조 검증)",
    "domain": "domain (값 의미 검증)",
    "state": "state (상태 검증)",
    "runtime": "runtime (안정성 검증)",
}

_PROFILE_KO = {
    "raw_image": "raw_image",
    "face_operation": "face_operation",
    "match_verdict": "match_verdict",
    "default": "default",
}

_EP_DESC = {
    "/health": "서버 상태 확인",
    "/api/v2/detect": "얼굴 인식/검출",
    "/api/v2/compare": "얼굴 비교",
    "/api/v2/identify": "얼굴 식별",
    "/api/v2/enroll": "얼굴 등록",
    "/api/v2/delete": "얼굴 데이터 삭제",
    "/api/v2/list": "등록 목록 조회",
    "/api/v2/get_config": "설정값 조회",
    "/api/v2/save_config": "설정값 저장",
    "/api/v2/reset_config": "설정 초기화",
    "/api/v2/get_license": "라이선스 조회",
    "/api/v2/get_version": "버전 정보 조회",
    "/api/v2/match": "템플릿 매칭",
    "/api/v2/verify": "템플릿 검증",
    "/api/v2/match-images": "이미지 매칭",
    "/api/v2/verify-template": "템플릿 검증(서버 템플릿)",
}

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _map_qfe_error_code(error_code: Any, msg: str, path: str = "") -> str | None:
    try:
        ec = int(error_code)
    except Exception:
        return None

    msg_l = (msg or "").lower()
    path = path or ""

    # 1) state — DB/리소스/설정 없음
    _STATE = {
        -20: "DATABASE_NOT_EXIST",
        -21: "FAILED_FILE_ALREADY_EXIST",
        -22: "FAILED_CREATE_DATABASE_FILE",
        -23: "FAILED_CREATE_TABLE",
        -24: "FAILED_SET_OPTIONS",
        -25: "FAILED_INSERT_USER",
        -26: "FAILED_DELETE_USER",
        -27: "FAILED_CLEAR_DATABASE",
        -28: "USER_NOT_FOUND",
        -29: "DATABASE_NOT_LOADED",
        -30: "CONFIG_NOT_EXIST",
        -31: "FAILED_CONFIG_CREATE",
        -43: "TEMPLATE_NOT_FOUND",
        -91: "FILE_NOT_FOUND",
    }
    if ec in _STATE:
        return f"상태 미충족 — {_STATE[ec]}"

    # 2) domain — 알고리즘/비즈니스 로직 실패
    _DOMAIN = {
        -32: "SYS_PARAM_NOT_SUPPORT",
        -33: "SYS_PARAM_OUT_OF_RANGE",
        -34: "INVALID_USER_ID",
        -35: "INVALID_SUB_ID",
        -40: "MAX_TEMPLATE_LIMIT",
        -41: "ADD_TEMPLATE",
        -42: "UPDATE_TEMPLATE",
        -50: "ENROLL_DIFFERENT_FACE",
        -200: "FAILED_FACE_DETECT",
        -201: "INVALID_ROI_COORDINATE",
        -202: "ROI_OUT_OF_IMAGE_BOUNDARY",
        -300: "FAILED_CHECK_REAL_FACE",
        -400: "FAILED_ESTIMATE_HEAD_POSE",
        -401: "FAILED_OUT_OF_RANGE",
        -500: "FAILED_MASK_SEGMENTATION",
        -600: "FAILED_ESTIMATE_FACE_ATTR",
        -700: "FAILED_EXTRACT_TEMPLATE",
        -800: "FAILED_GET_THRESHOLD",
    }
    if ec in _DOMAIN:
        return f"도메인 실패 — {_DOMAIN[ec]}"

    # 3) request/schema
    if ec == -90:
        return "요청 오류 — INVALID_PARAMETER"

    # 4) system — SDK/라이선스/모델
    _SYSTEM = {
        -2: "MEM_ALLOC", 
        -3: "CAPTURE_FRAME", 
        -4: "INSTANCE_NOT_EXIST",
        -5: "LICENSE", 
        -6: "BUSY", 
        -7: "THREAD_INSUFFICIENT",
        -8: "INVALID_POINTER", 
        -9: "SDK_INSTANCE_BUSY",
        -10: "CAMERA_NOT_EXIST", 
        -11: "FAILED_SET_CAMERA", 
        -12: "FAILED_UNSET_CAMERA",
        -60: "LICENSE_EXPIRED", 
        -61: "LICENSE_SUSPENDED", 
        -62: "LICENSE_INVALID_PRODUCT",
        -63: "LICENSE_ACCESS_NO_PERMISSION", 
        -64: "LICENSE_TIME_SYNC",
        -65: "LICENSE_INVALID_KEY", 
        -66: "LICENSE_INVALID_FILE",
        -67: "LICENSE_MACHINE_CHANGED", 
        -68: "LICENSE_TIME_MODIFIED", 
        -69: "LICENSE_METADATA",
        -100: "FAILED_LOAD_MODEL", 
        -101: "FAILED_GPU_SET",
        -9999: "UNKNOWN",
    }
    if ec in _SYSTEM:
        return f"시스템 오류 — {_SYSTEM[ec]}"

    # 5) generic -1
    if ec == -1:
        if any(x in msg_l for x in ["required", "unmarshal", "json", "type", "invalid request"]):
            return "타입 오류 (Type Mismatch)"
        if any(x in msg_l for x in ["base64", "decode", "padding", "encoding"]):
            return "포맷 오류 (Encoding/Decode)"
        if "failed to detect face" in msg_l or "error code: -200" in msg_l:
            return "도메인 실패 — FAILED_FACE_DETECT"
        if path in ("/api/v2/verify-template", "/api/v2/verify") and "verification failed" in msg_l:
            return "상태 미충족 — USER_NOT_FOUND"
        if path == "/api/v2/delete" and "failed to delete user" in msg_l:
            return "상태 미충족 — USER_NOT_FOUND"
        if "failed to get user template" in msg_l or "template not found" in msg_l or "user not found" in msg_l:
            return "상태 미충족 — USER_NOT_FOUND"
        return "정상적인 실패 응답"

    return None

def classify_failure_cause_from_item(item: dict[str, Any]) -> str:
    outcome = str(item.get("outcome", "")).lower()
    axis = str(item.get("axis") or "")
    reason_code = str(item.get("reason_code") or "")
    error_detail = str(item.get("error_detail") or "")
    expected_result_type = str(item.get("expected_result_type") or "")
    path = str(item.get("request_path") or "")
    response_success = item.get("response_success")
    response_error_code = item.get("response_error_code")
    response_data_error_code = item.get("response_data_error_code")
    response_msg = str(item.get("response_msg") or "")
    msg = response_msg.lower()
    data_status = str(item.get("response_data_status") or "").lower()
    actual_status = str(item.get("actual_status") or "")

    exc_type = str(item.get("exception_type") or "")
    exc_msg = str(item.get("exception_message") or "")
    longrepr = str(item.get("longrepr") or "")
    blob = f"{exc_type} {exc_msg} {longrepr} {error_detail}".lower()

    # 0. PASS
    if outcome == "passed":
        # state-tolerant positive: success=false/error_code<0 였지만 허용된 케이스
        if (
            response_success is False
            and (reason_code == "precondition_not_met" or error_detail.startswith("state."))
        ):
            return "PASS (상태 미충족)"
        return "PASS"

    # 1. 인프라 / 런타임
    if item.get("server_crashed") or "crash_detected" in blob or actual_status.startswith("5"):
        return "서버 Crash (5xx)"

    if "connection" in blob or "refused" in blob or "timeout" in blob:
        return "서버 미응답"

    # 2. 생성/문법 오류
    if "indentationerror" in blob or "syntaxerror" in blob:
        return "TC 생성 오류 (문법/들여쓰기)"

    # 3. Crash Probe 우선 판정
    if "validation_gap" in blob:
        return "엔드포인트 버그 (Validation 미수행)"
    
    mapped = _map_qfe_error_code(response_error_code, response_msg, path)
    if mapped:
        return mapped

    mapped = _map_qfe_error_code(response_data_error_code, response_msg, path)
    if mapped:
        return mapped

    # 4. 상태 미충족 — msg/reason_code 기반 추가 분류
    if reason_code == "precondition_not_met" or error_detail.startswith("state."):
        return "상태 미충족 — USER_NOT_FOUND"
    if "failed to get user template" in msg or "template not found" in msg:
        return "상태 미충족 — TEMPLATE_NOT_FOUND"
    if "failed to delete user" in msg or "user not found" in msg:
        return "상태 미충족 — USER_NOT_FOUND"
    if "template extraction failed" in msg:
        return "상태 미충족 — TEMPLATE_NOT_FOUND"
    if (
        "verification failed" in msg
        and path in ("/api/v2/verify-template", "/api/v2/verify")
    ):
        return "상태 미충족 — USER_NOT_FOUND"

    # 5. response body 문구 기반 분류
    # 5-1. 필수값/누락
    if "missing" in msg or "required" in msg:
        return "입력 누락 (Required Missing)"

    # 5-2. 타입/직렬화 오류
    if (
        "cannot unmarshal" in msg
        or "type mismatch" in msg
        or "invalid type" in msg
        or "wrong type" in msg
    ):
        return "타입 오류 (Type Mismatch)"

    # 5-3. 포맷/인코딩 오류
    if (
        "base64" in msg
        or "decode" in msg
        or "padding" in msg
        or "encoding" in msg
        or "invalid request" in msg and "json" in msg
    ):
        return "포맷 오류 (Encoding/Decode)"

    # 5-4. 범위 오류
    if "range" in msg or "out of range" in msg or "must be between" in msg:
        return "범위 오류 (Out of Range)"

    # 5-5. 리소스 없음
    if "not found" in msg or "no such" in msg or "does not exist" in msg:
        return "리소스 없음 (Not Found)"

    # 5-6. 중복
    if "duplicate" in msg or "already exists" in msg:
        return "중복 요청 (Duplicate)"

    # 5-7. threshold / match / verify 도메인 실패
    if "threshold" in msg:
        return "도메인 조건 실패 (Threshold/Match)"


    # 7. expected_result_type + success 조합
    if expected_result_type == "probe_only":
        if "assert" in blob or "failed:" in blob:
            return "TC 실패 (단언 오류)"
        return "Probe Only"

    if expected_result_type == "expected_fail":
        if response_success is True:
            if axis == "schema":
                return "엔드포인트 버그 (Validation 미수행)"
            return "엔드포인트 버그 (도메인 검증 미수행)"
        if response_success is False:
            return "정상적인 실패 응답"

    if expected_result_type == "expected_pass":
        if response_success is False:
            return "예상치 못한 실패"

    # 8. data.status 기반 보정
    if data_status in {"fail", "success"}:
        if expected_result_type == "expected_fail" and data_status == "success":
            return "엔드포인트 버그 (도메인 검증 미수행)"
        if expected_result_type == "expected_pass" and data_status == "fail":
            return "도메인 조건 실패 (Threshold/Match)"

    # 9. assertion/pytest failure
    if "assert" in blob or "failed:" in blob:
        return "TC 실패 (단언 오류)"

    return "알 수 없음"

class ExcelReportBuilder2:
    """excel_reporter.py 인터페이스 호환, 4-시트 한국어 워크북 생성."""

    def __init__(self, output_path: str | Path):
        self.output_path = Path(output_path)

    def build(
        self,
        runner_summary: dict[str, Any],
        pytest_json_path: str | Path | None = None,
        source_file: str = "",
        base_url: str = "",
        endpoints: list[dict] | None = None,
        allure_results_dir: str | Path | None = None,
        crash_probe_json_path: str | Path | None = None,
    ) -> Path:
        tests = self._load_test_results(
            pytest_json_path=pytest_json_path,
            allure_results_dir=allure_results_dir,
        )
        probe_tests = self._load_probe_tests(crash_probe_json_path)

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "TC 명세서"
        self._build_tc_spec(ws1, tests)

        ws2 = wb.create_sheet("엔드포인트별 요약")
        self._build_endpoint_summary(ws2, tests)

        ws3 = wb.create_sheet("종합 요약")
        self._build_overall_summary(ws3, runner_summary, tests, source_file, base_url,
                                    probe_tests=probe_tests)

        ws4 = wb.create_sheet("전체 결과 상세")
        self._build_detail_table(ws4, tests, base_url)

        ws5 = wb.create_sheet("Crash Probe")
        self._build_crash_probe_sheet(ws5, probe_tests)

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_path)
        return self.output_path

    def _build_tc_spec(self, ws, tests: list[dict[str, Any]]) -> None:
        self._title_banner(ws, "A1:Q1", "TC 명세서 (현재 Rule/Diag 반영)")
        ws.row_dimensions[1].height = 26

        headers = [
            "TC ID", "HTTP 메서드", "엔드포인트", "기능설명",
            "테스트유형", "세부유형", "프로파일",
            "기대결과유형", "Semantic Tag", "Policy",
            "테스트조건", "기댓값", "실행결과",
            "Data Error", "Match Score", "Match Status", "P/F",
        ]
        self._header_row(ws, 2, headers)

        col_widths = [10, 12, 32, 18, 18, 18, 16, 18, 16, 12, 44, 28, 34, 12, 14, 14, 8]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ca = Alignment(horizontal="center", vertical="top", wrap_text=True)
        la = Alignment(horizontal="left", vertical="top", wrap_text=True)

        for idx, item in enumerate(tests, start=1):
            info = self._parse_nodeid(item.get("nodeid", ""), {})
            method = (item.get("request_method") or info["method"]).upper()
            path = item.get("request_path") or info["path"] or ""
            rtype = item.get("rule_type") or info["rule_type"] or ""
            subtype = item.get("rule_subtype", "")
            profile = item.get("endpoint_profile", "")
            expected_result_type = item.get("expected_result_type", "")
            semantic_tag = item.get("semantic_tag", "")
            policy = item.get("policy", "")
            cond = item.get("condition") or info["condition"] or ""
            outcome = str(item.get("outcome", "")).lower()
            pf = "PASS" if outcome == "passed" else "FAIL"
            expected_display = item.get("expected_status_display") or info.get("expected_status", "")
            actual = self._build_actual_short(item)

            row_vals = [
                f"TC-{idx:04d}",
                method,
                path,
                _EP_DESC.get(path, ""),
                _TYPE_KO.get(rtype, rtype),
                subtype,
                _PROFILE_KO.get(profile, profile),
                _EXPECTED_RESULT_TYPE_KO.get(expected_result_type, expected_result_type),
                semantic_tag,
                policy,
                cond,
                expected_display,
                actual,
                item.get("response_data_error_code", ""),
                item.get("response_data_match_score", ""),
                item.get("response_data_status", ""),
                pf,
            ]

            r = idx + 2
            for ci, val in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=ci, value=val)
                c.border = _BORDER
                c.alignment = ca if ci in (1, 2, 14, 15, 16, 17) else la

            pf_cell = ws.cell(row=r, column=17)
            if pf == "PASS":
                pf_cell.font = Font(name="Arial", size=9, bold=True, color=_GREEN_DARK)
                pf_cell.fill = PatternFill("solid", start_color=_GREEN_FILL, end_color=_GREEN_FILL)
            else:
                pf_cell.font = Font(name="Arial", size=9, bold=True, color=_RED_FONT)
                pf_cell.fill = PatternFill("solid", start_color=_RED_FILL, end_color=_RED_FILL)

            if pf == "FAIL":
                for ci in range(1, 17):
                    ws.cell(row=r, column=ci).fill = PatternFill("solid", start_color="FFF0ED", end_color="FFF0ED")
            elif idx % 2 == 0:
                for ci in range(1, 17):
                    ws.cell(row=r, column=ci).fill = PatternFill("solid", start_color=_STRIPE, end_color=_STRIPE)

            ws.row_dimensions[r].height = max(18, min(72, 18 + len(cond) // 4))

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:Q{len(tests) + 2}"

    def _build_endpoint_summary(self, ws, tests: list[dict[str, Any]]) -> None:
        self._title_banner(ws, "A1:J1", "엔드포인트별 요약 (성격별 분리)")
        ws.row_dimensions[1].height = 26

        headers = [
            "HTTP 메서드", "엔드포인트", "기능설명",
            "전체 TC", "PASS", "FAIL",
            "Expected Pass 실패", "Expected Fail 이상동작", "Probe Only 실패", "주요 이슈",
        ]
        self._header_row(ws, 2, headers)

        col_widths = [12, 32, 18, 10, 10, 10, 16, 18, 16, 50]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ep_map: dict[tuple[str, str], dict[str, Any]] = {}
        for item in tests:
            info = self._parse_nodeid(item.get("nodeid", ""), {})
            method = (item.get("request_method") or info["method"]).upper() or "POST"
            path = item.get("request_path") or info["path"] or "(unknown)"
            key = (method, path)
            if key not in ep_map:
                ep_map[key] = {
                    "total": 0, "pass": 0, "fail": 0,
                    "expected_pass_fail": 0,
                    "expected_fail_misbehave": 0,
                    "probe_only_fail": 0,
                    "issues": [],
                }

            ep_map[key]["total"] += 1
            outcome = str(item.get("outcome", "")).lower()
            expected_result_type = str(item.get("expected_result_type") or "")
            failure_cause = classify_failure_cause_from_item(item)

            if outcome == "passed":
                ep_map[key]["pass"] += 1
            else:
                ep_map[key]["fail"] += 1
                if expected_result_type == "expected_pass":
                    ep_map[key]["expected_pass_fail"] += 1
                elif expected_result_type == "expected_fail":
                    ep_map[key]["expected_fail_misbehave"] += 1
                elif expected_result_type == "probe_only":
                    ep_map[key]["probe_only_fail"] += 1

                if failure_cause and failure_cause not in ep_map[key]["issues"]:
                    ep_map[key]["issues"].append(failure_cause)

        ca = Alignment(horizontal="center", vertical="center")
        la = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for ri, ((method, path), stat) in enumerate(sorted(ep_map.items(), key=lambda x: x[0][1]), start=3):
            issues = "; ".join(stat["issues"][:3]) or "없음"
            row_vals = [
                method, path, _EP_DESC.get(path, ""),
                stat["total"], stat["pass"], stat["fail"],
                stat["expected_pass_fail"], stat["expected_fail_misbehave"],
                stat["probe_only_fail"], issues,
            ]
            bg = _METHOD_COLORS.get(method, _WHITE)
            for ci, val in enumerate(row_vals, start=1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = _BORDER
                c.alignment = ca if ci in (1, 4, 5, 6, 7, 8, 9) else la
                if ci <= 3:
                    c.fill = PatternFill("solid", start_color=bg, end_color=bg)

            ws.cell(row=ri, column=5).fill = PatternFill("solid", start_color=_GREEN_FILL, end_color=_GREEN_FILL)
            if stat["fail"] > 0:
                ws.cell(row=ri, column=6).fill = PatternFill("solid", start_color=_RED_FILL, end_color=_RED_FILL)

        ws.freeze_panes = "A3"

    def _build_overall_summary(self, ws, summary: dict[str, Any], tests: list[dict[str, Any]], source_file: str, base_url: str, probe_tests: list[dict[str, Any]] | None = None) -> None:
        self._title_banner(ws, "A1:F1", "자동화 테스트 최종 리포트 — 종합 요약")
        ws.row_dimensions[1].height = 28
        for col, width in zip(["A", "B", "C", "D", "E", "F"], [22, 18, 18, 18, 18, 36]):
            ws.column_dimensions[col].width = width

        total = len(tests)
        passed = sum(1 for t in tests if str(t.get("outcome", "")).lower() == "passed")
        failed = total - passed
        rate = round(passed / total * 100, 1) if total else 0.0

        expected_pass_fail = 0
        expected_fail_misbehave = 0
        probe_only_fail = 0
        precondition_fail = 0

        for item in tests:
            if str(item.get("outcome", "")).lower() == "passed":
                continue
            expected_result_type = str(item.get("expected_result_type") or "")
            reason_code = str(item.get("reason_code") or "")
            response_success = item.get("response_success")
            if expected_result_type == "expected_pass":
                expected_pass_fail += 1
                if reason_code == "precondition_not_met":
                    precondition_fail += 1
            elif expected_result_type == "expected_fail":
                if response_success is True:
                    expected_fail_misbehave += 1
            elif expected_result_type == "probe_only":
                probe_only_fail += 1

        self._section_header(ws, "A3:F3", "📊 핵심 지표 (KPI)")
        kpi_labels = ["총 TC 수", "PASS", "FAIL", "통과율"]
        kpi_values = [str(total), str(passed), str(failed), f"{rate}%"]
        kpi_fgs = [_BLUE_LIGHT, _GREEN_FILL, _RED_FILL, _GREEN_FILL if rate >= 80 else _YELLOW_FILL if rate >= 60 else _RED_FILL]
        kpi_fonts = [_BLUE_DARK, _GREEN_DARK, _RED_FONT, _GREEN_DARK if rate >= 80 else "7F6000" if rate >= 60 else _RED_FONT]

        for ci, (label, value, fg, fc) in enumerate(zip(kpi_labels, kpi_values, kpi_fgs, kpi_fonts), start=1):
            lc = ws.cell(row=4, column=ci, value=label)
            lc.font = Font(bold=True, size=10, color=_WHITE)
            lc.fill = PatternFill("solid", start_color=_BLUE_DARK, end_color=_BLUE_DARK)
            lc.alignment = Alignment(horizontal="center", vertical="center")
            lc.border = _BORDER
            vc = ws.cell(row=5, column=ci, value=value)
            vc.font = Font(bold=True, size=16, color=fc)
            vc.fill = PatternFill("solid", start_color=fg, end_color=fg)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border = _BORDER

        self._section_header(ws, "A7:F7", "🧭 현재 Rule/Oracle 구조 기반 지표")
        extra_rows = [
            ("Expected Pass 실패", expected_pass_fail, _RED_FILL, _RED_FONT),
            ("Expected Fail 이상동작", expected_fail_misbehave, _RED_FILL, _RED_FONT),
            ("Probe Only 실패", probe_only_fail, _YELLOW_FILL, "7F6000"),
            ("Precondition 실패", precondition_fail, "FCE5CD", "9C5700"),
        ]
        for i, (label, value, fill, font_color) in enumerate(extra_rows, start=8):
            lc = ws.cell(row=i, column=1, value=label)
            lc.font = Font(bold=True)
            lc.fill = PatternFill("solid", start_color=_GRAY_HEADER, end_color=_GRAY_HEADER)
            lc.alignment = Alignment(horizontal="left", vertical="center")
            lc.border = _BORDER
            vc = ws.cell(row=i, column=2, value=value)
            vc.font = Font(bold=True, color=font_color)
            vc.fill = PatternFill("solid", start_color=fill, end_color=fill)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border = _BORDER
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)

        row_cursor = 13
        self._section_header(ws, f"A{row_cursor}:F{row_cursor}", "🔍 테스트유형별 분석")
        type_headers = ["테스트유형", "전체", "PASS", "FAIL", "통과율", "판정"]
        for ci, h in enumerate(type_headers, start=1):
            c = ws.cell(row=row_cursor + 1, column=ci, value=h)
            c.font = Font(bold=True, color=_WHITE)
            c.fill = PatternFill("solid", start_color=_BLUE_DARK, end_color=_BLUE_DARK)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _BORDER

        type_counts: dict[str, dict[str, int]] = {}
        for item in tests:
            info = self._parse_nodeid(item.get("nodeid", ""), {})
            rtype = item.get("rule_type") or info["rule_type"] or "기타"
            type_counts.setdefault(rtype, {"total": 0, "pass": 0, "fail": 0})
            type_counts[rtype]["total"] += 1
            if str(item.get("outcome", "")).lower() == "passed":
                type_counts[rtype]["pass"] += 1
            else:
                type_counts[rtype]["fail"] += 1

        row_cursor += 2
        for rtype, stat in sorted(type_counts.items(), key=lambda x: -x[1]["total"]):
            t = stat["total"]
            p = stat["pass"]
            f = stat["fail"]
            r2 = round(p / t * 100, 1) if t else 0.0
            verdict = "✅ 양호" if r2 >= 90 else "⚠️ 주의" if r2 >= 60 else "❌ 불량"
            row_vals = [_TYPE_KO.get(rtype, rtype), t, p, f, f"{r2}%", verdict]
            for ci, val in enumerate(row_vals, start=1):
                c = ws.cell(row=row_cursor, column=ci, value=val)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = _BORDER
                if ci == 1:
                    c.alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=row_cursor, column=3).fill = PatternFill("solid", start_color=_GREEN_FILL, end_color=_GREEN_FILL)
            if f > 0:
                ws.cell(row=row_cursor, column=4).fill = PatternFill("solid", start_color=_RED_FILL, end_color=_RED_FILL)
            row_cursor += 1

        self._section_header(ws, f"A{row_cursor + 1}:F{row_cursor + 1}", "⚠️ 주요 실패 원인 분석")
        fail_reasons: dict[str, int] = {}
        for item in tests:
            if str(item.get("outcome", "")).lower() == "passed":
                continue
            cause = classify_failure_cause_from_item(item)
            fail_reasons[cause] = fail_reasons.get(cause, 0) + 1

        top_reasons = sorted(fail_reasons.items(), key=lambda x: -x[1])[:10]
        for i, (reason, cnt) in enumerate(top_reasons, start=1):
            rr = row_cursor + 1 + i
            ws.cell(row=rr, column=1, value=f"{i}.").alignment = Alignment(horizontal="center")
            ws.cell(row=rr, column=1).border = _BORDER
            c = ws.cell(row=rr, column=2, value=reason)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = _BORDER
            ws.merge_cells(f"B{rr}:E{rr}")
            cnt_c = ws.cell(row=rr, column=6, value=f"{cnt}건")
            cnt_c.alignment = Alignment(horizontal="center")
            cnt_c.fill = PatternFill("solid", start_color=_RED_FILL, end_color=_RED_FILL)
            cnt_c.border = _BORDER

        row_cursor = row_cursor + 2 + len(top_reasons)

        # ── Robustness Layer (Crash Probe) 요약 ─────────────────────────────
        if probe_tests:
            self._section_header(ws, f"A{row_cursor}:F{row_cursor}", "🛡️ Robustness Layer (B) — Crash Probe 요약")
            probe_hdr = ["전체 Probe", "CRASH_DETECTED", "VALIDATION_GAP", "GRACEFUL_REJECTION", "기타"]
            crash_c    = sum(1 for t in probe_tests if self._classify_probe(t) == "CRASH_DETECTED")
            gap_c      = sum(1 for t in probe_tests if self._classify_probe(t) == "VALIDATION_GAP")
            graceful_c = sum(1 for t in probe_tests if self._classify_probe(t) == "GRACEFUL_REJECTION")
            other_c    = len(probe_tests) - crash_c - gap_c - graceful_c
            probe_vals = [len(probe_tests), crash_c, gap_c, graceful_c, other_c]
            probe_bgs  = ["BDD7EE", "F4CCCC", "FCE5CD", "C6EFCE", "EFEFEF"]
            probe_fcs  = ["1F497D", "C00000", "7F3F00", "375623", "404040"]

            for ci, (label, value, bg, fc) in enumerate(zip(probe_hdr, probe_vals, probe_bgs, probe_fcs), start=1):
                lc = ws.cell(row=row_cursor + 1, column=ci, value=label)
                lc.font = Font(bold=True, color=_WHITE, size=9)
                lc.fill = PatternFill("solid", start_color=_BLUE_DARK, end_color=_BLUE_DARK)
                lc.alignment = Alignment(horizontal="center", vertical="center")
                lc.border = _BORDER
                vc = ws.cell(row=row_cursor + 2, column=ci, value=value)
                vc.font = Font(bold=True, size=14, color=fc)
                vc.fill = PatternFill("solid", start_color=bg, end_color=bg)
                vc.alignment = Alignment(horizontal="center", vertical="center")
                vc.border = _BORDER

            # Crash probe 판정
            verdict_row = row_cursor + 3
            verdict_msg = (
                f"⛔ 서버 크래시 {crash_c}건 감지 — CGO 호출 전 입력 검증 필요"
                if crash_c > 0
                else ("⚠️ VALIDATION_GAP 감지 — 비정상 입력 수락 여부 확인 필요" if gap_c > 0 else "✅ 모든 프로브 통과 (GRACEFUL_REJECTION)")
            )
            vbg = "F4CCCC" if crash_c > 0 else ("FCE5CD" if gap_c > 0 else "C6EFCE")
            vc = ws.cell(row=verdict_row, column=1, value=verdict_msg)
            vc.font = Font(bold=True, size=10,
                           color=("C00000" if crash_c > 0 else ("7F3F00" if gap_c > 0 else "375623")))
            vc.fill = PatternFill("solid", start_color=vbg, end_color=vbg)
            vc.alignment = Alignment(horizontal="left", vertical="center")
            vc.border = _BORDER
            ws.merge_cells(f"A{verdict_row}:F{verdict_row}")

            row_cursor = verdict_row + 2

        self._section_header(ws, f"A{row_cursor}:F{row_cursor}", "🖥️ 실행 환경 정보")
        env_rows = [
            ("생성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("소스 파일", source_file),
            ("서버 URL", base_url),
            ("총 실행 시간", f"{summary.get('duration_seconds', '')} 초"),
            ("테스트 프레임워크", "pytest + pytest-json-report"),
            ("리포트 작성", "excel_reporter2.py (AutoTC)"),
        ]
        for i, (label, value) in enumerate(env_rows, start=1):
            rr = row_cursor + i
            lc = ws.cell(row=rr, column=1, value=label)
            lc.font = Font(bold=True)
            lc.fill = PatternFill("solid", start_color=_GRAY_HEADER, end_color=_GRAY_HEADER)
            lc.border = _BORDER
            vc = ws.cell(row=rr, column=2, value=value)
            vc.border = _BORDER
            vc.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f"B{rr}:F{rr}")

    def _build_detail_table(self, ws, tests: list[dict[str, Any]], base_url: str) -> None:
        self._title_banner(ws, "A1:R1", "자동화 테스트 최종 리포트 — 전체 결과 상세")
        ws.row_dimensions[1].height = 26

        headers = [
            "TC ID",
            "엔드포인트",
            "테스트유형",
            "세부유형",
            "프로파일",
            "예상결과",
            "Axis",
            "Reason Code",
            "테스트조건",
            "예상 HTTP",
            "예상 응답",
            "실제 HTTP",
            "실제 응답",
            "Data Error",
            "Match Score",
            "Match Status",
            "P/F",
            "소요시간(s)",
        ]
        self._header_row(ws, 2, headers)

        col_widths = [
            10, 42, 18, 18, 16, 16,
            18, 20, 46, 16, 38,
            12, 38, 12, 14, 14,
            8, 12,
        ]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ca = Alignment(horizontal="center", vertical="top", wrap_text=True)
        la = Alignment(horizontal="left", vertical="top", wrap_text=True)

        def split_expected(item: dict[str, Any], fallback: str) -> tuple[str, str]:
            expected_http = str(item.get("expected_http") or "").strip()
            expected_app = str(item.get("expected_app") or "").strip()

            if expected_http and expected_app:
                return expected_http, expected_app

            display = str(item.get("expected_status_display") or fallback or "").strip()
            if not display:
                return "", ""

            # expected_http가 200/400/422/503 형태일 수 있으므로
            # 반드시 공백 포함 구분자 " / " 기준으로만 분리
            if " / " in display:
                left, right = display.split(" / ", 1)
                return left.strip(), right.strip()

            m = re.match(r"^(\d{3}(?:/\d{3})*)\s*(.*)$", display)
            if m:
                return m.group(1).strip(), m.group(2).strip()

            return display, ""

        def excel_linebreak(value: Any) -> str:
            """
            Excel 셀 내부 Alt+Enter 효과.
            openpyxl에서는 문자열 안의 \\n + wrap_text=True면 줄바꿈 표시됨.
            단, HTTP 후보값 200/400/422 같은 슬래시는 깨지 않기 위해
            공백 포함 ' / ' 만 줄바꿈으로 바꾼다.
            """
            if value is None:
                return ""
            text = str(value)
            text = text.replace(" / ", "\n")
            return text

        for idx, item in enumerate(tests, start=1):
            info = self._parse_nodeid(item.get("nodeid", ""), {})

            method = (item.get("request_method") or info["method"] or "").upper()
            path = item.get("request_path") or info["path"] or ""
            endpoint = f"{method} {path}".strip()

            rtype = item.get("rule_type") or info["rule_type"] or ""
            subtype = item.get("rule_subtype", "")
            profile = item.get("endpoint_profile", "")
            expected_result_type = item.get("expected_result_type", "")

            axis = _AXIS_LABEL.get(item.get("axis", ""), item.get("axis", ""))
            reason_code = item.get("reason_code", "")
            cond = item.get("condition") or info["condition"] or ""

            pf = "PASS" if str(item.get("outcome", "")).lower() == "passed" else "FAIL"
            dur = round(float(item.get("duration", 0) or 0), 3)

            expected_display = item.get("expected_status_display") or info.get("expected_status", "")
            expected_http, expected_resp = split_expected(item, expected_display)

            act_http = str(item.get("actual_status") or "")
            act_resp = self._build_actual_resp(item)

            row_vals = [
                f"TC-{idx:04d}",
                endpoint,
                _TYPE_KO.get(rtype, rtype),
                subtype,
                _PROFILE_KO.get(profile, profile),
                _EXPECTED_RESULT_TYPE_KO.get(expected_result_type, expected_result_type),
                axis,
                reason_code,
                cond,
                expected_http,
                excel_linebreak(expected_resp),
                act_http,
                excel_linebreak(act_resp),
                item.get("response_data_error_code", ""),
                item.get("response_data_match_score", ""),
                item.get("response_data_status", ""),
                pf,
                dur,
            ]

            r = idx + 2
            for ci, val in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=ci, value=val)
                c.border = _BORDER
                c.alignment = ca if ci in (1, 3, 4, 5, 6, 10, 12, 14, 15, 16, 17, 18) else la

            pf_c = ws.cell(row=r, column=17)
            if pf == "PASS":
                pf_c.font = Font(bold=True, color=_GREEN_DARK)
                pf_c.fill = PatternFill("solid", start_color=_GREEN_FILL, end_color=_GREEN_FILL)
            else:
                pf_c.font = Font(bold=True, color=_RED_FONT)
                pf_c.fill = PatternFill("solid", start_color=_RED_FILL, end_color=_RED_FILL)

            if pf == "FAIL":
                for ci in range(1, len(headers) + 1):
                    ws.cell(row=r, column=ci).fill = PatternFill(
                        "solid",
                        start_color="FFF0ED",
                        end_color="FFF0ED",
                    )
                pf_c.fill = PatternFill("solid", start_color=_RED_FILL, end_color=_RED_FILL)

            elif idx % 2 == 0:
                for ci in range(1, len(headers) + 1):
                    ws.cell(row=r, column=ci).fill = PatternFill(
                        "solid",
                        start_color=_STRIPE,
                        end_color=_STRIPE,
                    )
                pf_c.fill = PatternFill("solid", start_color=_GREEN_FILL, end_color=_GREEN_FILL)

            # 응답 필드에 줄바꿈이 들어가면 행 높이를 조금 키움
            if "\n" in str(row_vals[10]) or "\n" in str(row_vals[12]):
                ws.row_dimensions[r].height = 52
            else:
                ws.row_dimensions[r].height = 34

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:R{len(tests) + 2}"

    def _load_test_results(self, pytest_json_path: str | Path | None, allure_results_dir: str | Path | None) -> list[dict[str, Any]]:
        """테스트 결과를 로드한다.

        우선순위:
          1. report.json (pytest-json-report) — run-pipeline 아티팩트에서 다운로드된 최신 파일
          2. allure-results/ — report.json 이 없거나 전체가 인프라 실패(서버 미구동)인 경우 fallback

        pytest_report.json 은 구 버전 파이프라인 잔재이므로 절대 사용하지 않는다.
        generate-report job 이 pytest-raw 아티팩트를 먼저 다운로드하므로
        report.json 은 항상 현재 실행 결과를 담고 있다.
        """
        tests: list[dict[str, Any]] = []
        if pytest_json_path:
            p = Path(pytest_json_path)
            if p.exists() and p.is_file():
                try:
                    raw = json.loads(p.read_text(encoding="utf-8"))
                    if isinstance(raw, dict) and "tests" in raw:
                        tests = self._normalize_pytest_json(raw)
                except Exception:
                    pass

        # allure-results fallback:
        #   - report.json 이 없거나 테스트가 0건인 경우
        #   - 또는 모든 테스트가 서버 미구동(ConnectionError)으로 인한 인프라 실패인 경우
        #   (allure-results 는 --alluredir 옵션으로 매 실행 전 초기화 후 생성됨)
        if allure_results_dir and self._should_fallback_to_allure(tests):
            d = Path(allure_results_dir)
            if d.exists() and d.is_dir():
                allure_tests = self._normalize_allure(d)
                if allure_tests:
                    print(f"[ExcelReporter2] allure fallback: {len(tests)} pytest → {len(allure_tests)} allure")
                    tests = allure_tests

        diag_map = self._load_diag_jsonl(pytest_json_path)
        if diag_map:
            for t in tests:
                diag = diag_map.get(t.get("nodeid", ""))
                if diag:
                    self._apply_diag(t, diag)

        return tests

    @staticmethod
    def _should_fallback_to_allure(tests: list[dict[str, Any]]) -> bool:
        """report.json 결과가 없거나 전부 인프라 실패(서버 미구동)이면 True."""
        if not tests:
            return True
        infra_types = {"ConnectionError", "ConnectionRefusedError", "NewConnectionError", "MaxRetryError"}
        return all(
            (t.get("exception_type") or "") in infra_types
            or "connection_refused" in (t.get("reason_code") or t.get("error_detail") or "").lower()
            for t in tests
        )

    def _normalize_pytest_json(self, raw: dict) -> list[dict[str, Any]]:
        out = []
        for t in raw.get("tests", []):
            call = t.get("call") or {}
            meta = self._extract_tc_meta(t)
            inline_diag: dict[str, Any] = meta.pop("_diag_from_up", {}) or {}
            item: dict[str, Any] = {
                "nodeid": t.get("nodeid", ""),
                "outcome": t.get("outcome", "unknown"),
                "duration": call.get("duration", t.get("duration", 0)),
                "longrepr": str(call.get("longrepr") or t.get("longrepr") or ""),
                "rule_type": meta.get("rule_type", ""),
                "rule_subtype": meta.get("rule_subtype", ""),
                "endpoint_profile": meta.get("endpoint_profile", ""),
                "semantic_tag": meta.get("semantic_tag", ""),
                "policy": meta.get("policy", ""),
                "expected_result_type": meta.get("expected_result_type", ""),
                "target_param": meta.get("target_param", ""),
                "condition": meta.get("condition", ""),
                "request_method": meta.get("request_method", ""),
                "request_path": meta.get("request_path", ""),
                "request_query": meta.get("request_query", {}),
                "request_headers": meta.get("request_headers", {}),
                "request_body": meta.get("request_body"),
                "expected_status": meta.get("expected_status", []),
                "expected_status_display": self._coerce_expected(meta),
                "actual_status": meta.get("actual_status", ""),
                "response_text": meta.get("response_text", ""),
                "exception_type": meta.get("exception_type", ""),
                "exception_message": meta.get("exception_message", ""),
                "server_crashed": meta.get("server_crashed", False),
                "axis": "",
                "reason_code": "",
                "error_detail": "",
                "response_success": None,
                "response_error_code": None,
                "response_msg": None,
                "response_data": None,
                "response_data_error_code": None,
                "response_data_match_score": None,
                "response_data_status": None,
            }
            if inline_diag:
                self._apply_diag(item, inline_diag)
            out.append(item)
        return out

    def _normalize_allure(self, d: Path) -> list[dict[str, Any]]:
        out = []
        for p in sorted(d.glob("*-result.json")):
            try:
                raw = json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                continue
            start = raw.get("start")
            stop = raw.get("stop")
            dur = round((stop - start) / 1000.0, 3) if isinstance(start, (int, float)) and isinstance(stop, (int, float)) and stop >= start else 0.0
            details = raw.get("statusDetails") or {}
            out.append({
                "nodeid": raw.get("fullName") or raw.get("name") or p.stem,
                "outcome": str(raw.get("status", "unknown")).lower(),
                "duration": dur,
                "longrepr": str(details.get("message") or details.get("trace") or ""),
                "rule_type": "", "rule_subtype": "", "endpoint_profile": "", "semantic_tag": "", "policy": "", "expected_result_type": "",
                "target_param": "", "condition": "", "request_method": "", "request_path": "", "actual_status": "", "response_text": "",
                "exception_type": "", "exception_message": "", "error_detail": "", "axis": "", "reason_code": "",
                "request_query": {}, "request_headers": {}, "request_body": None, "expected_status": [], "expected_status_display": "",
                "server_crashed": False, "response_success": None, "response_error_code": None, "response_msg": None,
                "response_data": None, "response_data_error_code": None, "response_data_match_score": None, "response_data_status": None,
            })
        return out

    @staticmethod
    def _load_diag_jsonl(pytest_json_path: str | Path | None) -> dict[str, dict]:
        candidates = []
        if pytest_json_path:
            candidates.append(Path(pytest_json_path).parent / "test_diag.jsonl")
        candidates.append(Path("reports/test_diag.jsonl"))
        for p in candidates:
            if not p.exists():
                continue
            result: dict[str, dict] = {}
            try:
                for line in p.read_text(encoding="utf-8").splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    rec = json.loads(line)
                    test_id = rec.get("test_id", "")
                    diag = rec.get("diag")
                    if test_id and isinstance(diag, dict):
                        result[test_id] = diag
                return result
            except Exception:
                pass
        return {}

    @staticmethod
    def _apply_diag(item: dict[str, Any], diag: dict[str, Any]) -> None:
        item["axis"] = diag.get("axis", "")
        item["reason_code"] = diag.get("reason_code", "")
        item["target_param"] = diag.get("target_field", item.get("target_param", ""))
        item["condition"] = diag.get("test_condition", item.get("condition", ""))
        
        exp_http = diag.get("expected_http") or ""
        exp_app = diag.get("expected_app") or ""
        parts = [p for p in [exp_http, exp_app] if p]
        if parts:
            item["expected_status_display"] = " / ".join(parts)
        if diag.get("actual_status") is not None:
            item["actual_status"] = str(diag["actual_status"])
        for k in ("request_body", "request_query", "request_headers"):
            v = diag.get(k)
            if v is not None:
                item[k] = v
        item["response_text"] = diag.get("response_snippet", item.get("response_text", ""))
        item["exception_type"] = diag.get("exception_type", "")
        item["exception_message"] = diag.get("exception_message", "")
        item["server_crashed"] = bool(diag.get("server_crash", False))
        item["error_detail"] = diag.get("error_detail", "")
        item["response_success"] = diag.get("response_success")
        item["response_error_code"] = diag.get("response_error_code")
        item["response_msg"] = diag.get("response_msg")
        item["response_data"] = diag.get("response_data")
        item["response_data_error_code"] = diag.get("response_data_error_code")
        item["response_data_match_score"] = diag.get("response_data_match_score")
        item["response_data_status"] = diag.get("response_data_status")
        item["response_data_verified"] = diag.get("response_data_verified")
        item["probe_endpoint"] = diag.get("probe_endpoint", item.get("probe_endpoint", ""))
        item["probe_label"] = diag.get("probe_label", item.get("probe_label", ""))
        item["probe_input"] = diag.get("probe_input", item.get("probe_input"))
        item["probe_severity"] = diag.get("probe_severity", item.get("probe_severity", ""))
        item["probe_classification"] = diag.get("probe_classification", item.get("probe_classification", ""))

    @staticmethod
    def _extract_tc_meta(test_obj: dict[str, Any]) -> dict[str, Any]:
        meta: dict[str, Any] = {}
        md = test_obj.get("metadata") or {}
        if isinstance(md, dict):
            tc_meta = md.get("tc_meta")
            if isinstance(tc_meta, dict):
                meta.update(tc_meta)

        for up in test_obj.get("user_properties", []) or []:
            if isinstance(up, dict):
                if "tc_meta" in up and isinstance(up["tc_meta"], dict):
                    meta.update(up["tc_meta"])
                if "diag" in up and isinstance(up["diag"], dict):
                    meta["_diag_from_up"] = up["diag"]

                if "probe_meta" in up and isinstance(up["probe_meta"], dict):
                    meta.update(up["probe_meta"])
                if "probe_diag" in up and isinstance(up["probe_diag"], dict):
                    meta["_probe_diag_from_up"] = up["probe_diag"]

            elif isinstance(up, (list, tuple)) and len(up) == 2:
                if up[0] == "tc_meta" and isinstance(up[1], dict):
                    meta.update(up[1])
                if up[0] == "diag" and isinstance(up[1], dict):
                    meta["_diag_from_up"] = up[1]

                if up[0] == "probe_meta" and isinstance(up[1], dict):
                    meta.update(up[1])
                if up[0] == "probe_diag" and isinstance(up[1], dict):
                    meta["_probe_diag_from_up"] = up[1]

        return meta

    @staticmethod
    def _coerce_expected(meta: dict[str, Any]) -> str:
        if meta.get("expected_status_display"):
            return str(meta["expected_status_display"])
        if meta.get("expected_status"):
            return ", ".join(map(str, meta["expected_status"]))
        return ""

    def _parse_nodeid(self, nodeid: str, src_cache: dict) -> dict[str, str]:
        result = {"method": "", "path": "", "rule_type": "", "condition": "", "expected_status": ""}
        if not nodeid:
            return result
        file_part, func_name = (nodeid.split("::", 1) if "::" in nodeid else ("", nodeid))
        if file_part:
            method, path = self._read_file_header(file_part, src_cache)
            result["method"] = method
            result["path"] = path
        if not result["method"]:
            m = re.search(r"test_(get|post|put|delete|patch)__", func_name, re.IGNORECASE)
            if m:
                result["method"] = m.group(1).upper()
        fn = func_name.lower()
        if fn.endswith("_positive"):
            result.update(rule_type="positive", condition="정상 요청 (모든 필수 필드 유효값 포함)", expected_status="HTTP 200 / success=true")
        elif m2 := re.search(r"_missing_body_(.+)$", fn):
            result.update(rule_type="missing_required", condition=f"필수 바디 필드 누락: {m2.group(1)}", expected_status="HTTP 200 / success=false")
        elif fn.endswith("_missing_body"):
            result.update(rule_type="missing_required", condition="요청 바디 전체 누락", expected_status="HTTP 200 / success=false")
        elif m2 := re.search(r"_missing_(.+)$", fn):
            result.update(rule_type="missing_required", condition=f"필수 파라미터 누락: {m2.group(1)}", expected_status="HTTP 200 / success=false")
        elif m2 := re.search(r"_wrong_type_body_(.+)$", fn):
            result.update(rule_type="wrong_type", condition=f"바디 필드 타입 오류: {m2.group(1)}", expected_status="HTTP 200 / success=false")
        elif m2 := re.search(r"_wrong_type_(.+)$", fn):
            result.update(rule_type="wrong_type", condition=f"파라미터 타입 오류: {m2.group(1)}", expected_status="HTTP 200 / success=false")
        elif m2 := re.search(r"_boundary_body_(.+?)_(.+)$", fn):
            result.update(rule_type="boundary", condition=f"경계값 탐색: {m2.group(1)} [{m2.group(2)}]", expected_status="HTTP 200 / 500 미만")
        elif m2 := re.search(r"_boundary_(.+?)_(.+)$", fn):
            result.update(rule_type="boundary", condition=f"경계값 탐색: {m2.group(1)} [{m2.group(2)}]", expected_status="HTTP 200 / 500 미만")
        elif m2 := re.search(r"_semantic_(.+?)_(.+)$", fn):
            result.update(rule_type="semantic_probe", condition=f"시맨틱 탐색: {m2.group(1)} [{m2.group(2)}]", expected_status="HTTP 200 / 500 미만")
        elif m2 := re.search(r"_invalid_enum_body_(.+)$", fn):
            result.update(rule_type="invalid_enum", condition=f"허용 범위 외 열거형: {m2.group(1)}", expected_status="HTTP 200 / success=false")
        return result

    def _read_file_header(self, file_path: str, cache: dict[str, str]) -> tuple[str, str]:
        if not file_path:
            return "", ""
        if file_path not in cache:
            p = Path(file_path)
            cache[file_path] = p.read_text(encoding="utf-8") if p.exists() else ""
        src = cache[file_path]
        m = re.search(r"^#\s*(GET|POST|PUT|DELETE|PATCH)\s+(\S+)", src, re.MULTILINE | re.IGNORECASE)
        if m:
            return m.group(1).upper(), m.group(2).strip()
        return "", ""

    def _build_actual_short(self, item: dict[str, Any]) -> str:
        parts = []
        sts = item.get("actual_status")
        if sts:
            parts.append(str(sts))
        rs = item.get("response_success")
        if rs is not None:
            parts.append(f"success={'true' if rs else 'false'}")
        ec = item.get("response_error_code")
        if ec not in (None, ""):
            parts.append(f"error_code={ec}")
        dec = item.get("response_data_error_code")
        if dec not in (None, ""):
            parts.append(f"data.error_code={dec}")
        ms = item.get("response_data_match_score")
        if ms not in (None, ""):
            parts.append(f"data.match_score={ms}")
        ds = item.get("response_data_status")
        if ds not in (None, ""):
            parts.append(f"data.status={ds}")
        msg = item.get("response_msg") or ""
        if msg:
            parts.append(str(msg)[:40])
        return " / ".join(parts)

    # ─── Crash Probe 지원 ─────────────────────────────────────────────────────

    _PROBE_CLS_COLORS: dict[str, str] = {
        "CRASH_DETECTED":     "F4CCCC",
        "VALIDATION_GAP":     "FCE5CD",
        "GRACEFUL_REJECTION": "C6EFCE",
        "SKIPPED":            "FFF2CC",
        "OTHER_FAILURE":      "EFEFEF",
    }

    def _normalize_probe_tests(self, raw: dict[str, Any]) -> list[dict[str, Any]]:
        out = []
        for t in raw.get("tests", []):
            call = t.get("call") or {}
            meta = self._extract_tc_meta(t)
            probe_diag = meta.pop("_probe_diag_from_up", {}) or {}

            item = {
                "nodeid": t.get("nodeid", ""),
                "outcome": t.get("outcome", "unknown"),
                "duration": call.get("duration", t.get("duration", 0)),
                "longrepr": str(call.get("longrepr") or t.get("longrepr") or ""),
                "target_param": meta.get("target_field", ""),
                "probe_endpoint": meta.get("probe_endpoint", ""),
                "probe_label": meta.get("probe_label", ""),
                "probe_input": meta.get("probe_input"),
                "probe_severity": meta.get("probe_severity", ""),
                "expected_result_type": meta.get("expected_result_type", "probe_only"),
                "actual_status": "",
                "response_success": None,
                "response_error_code": None,
                "response_msg": None,
                "error_detail": "",
                "axis": "runtime",
                "reason_code": "probe_runtime",
                "probe_classification": meta.get("probe_classification", ""),
            }

            if probe_diag:
                self._apply_diag(item, probe_diag)

            out.append(item)
        return out

    def _load_probe_tests(self, report_path: str | Path | None) -> list[dict[str, Any]]:
        if not report_path:
            return []
        p = Path(report_path)
        if not p.exists():
            return []
        try:
            raw = json.loads(p.read_text(encoding="utf-8"))
            return self._normalize_probe_tests(raw)
        except Exception:
            return []

    @staticmethod
    def _probe_failure_reason(t: dict, cls: str) -> str:
        """Crash Probe 행의 Failure Reason 컬럼 — 분류 기반 의미있는 설명."""
        ec = t.get("response_error_code")
        msg = str(t.get("response_msg") or "")
        outcome = str(t.get("outcome", "")).lower()
        _STATE = {
            -20: "DATABASE_NOT_EXIST", -21: "FAILED_FILE_ALREADY_EXIST",
            -22: "FAILED_CREATE_DATABASE_FILE", -23: "FAILED_CREATE_TABLE",
            -24: "FAILED_SET_OPTIONS", -25: "FAILED_INSERT_USER",
            -26: "FAILED_DELETE_USER", -27: "FAILED_CLEAR_DATABASE",
            -28: "USER_NOT_FOUND", -29: "DATABASE_NOT_LOADED",
            -30: "CONFIG_NOT_EXIST", -31: "FAILED_CONFIG_CREATE",
            -43: "TEMPLATE_NOT_FOUND", -91: "FILE_NOT_FOUND",
        }
        _DOMAIN = {
            -32: "SYS_PARAM_NOT_SUPPORT", -33: "SYS_PARAM_OUT_OF_RANGE",
            -34: "INVALID_USER_ID", -35: "INVALID_SUB_ID",
            -40: "MAX_TEMPLATE_LIMIT", -41: "ADD_TEMPLATE", -42: "UPDATE_TEMPLATE",
            -50: "ENROLL_DIFFERENT_FACE",
            -200: "FAILED_FACE_DETECT", -201: "INVALID_ROI_COORDINATE",
            -202: "ROI_OUT_OF_IMAGE_BOUNDARY", -300: "FAILED_CHECK_REAL_FACE",
            -400: "FAILED_ESTIMATE_HEAD_POSE", -401: "FAILED_OUT_OF_RANGE",
            -500: "FAILED_MASK_SEGMENTATION", -600: "FAILED_ESTIMATE_FACE_ATTR",
            -700: "FAILED_EXTRACT_TEMPLATE", -800: "FAILED_GET_THRESHOLD",
        }
        try:
            ec_int = int(ec)
            if ec_int in _STATE:
                return f"상태 미충족 — {_STATE[ec_int]}"
            if ec_int in _DOMAIN:
                return f"도메인 실패 — {_DOMAIN[ec_int]}"
            if ec_int == -90:
                return "요청 오류 — INVALID_PARAMETER"
            if ec_int == -1:
                return "정상적인 실패 응답 (ec=-1)"
        except (TypeError, ValueError):
            pass
        if cls == "CRASH_DETECTED":
            return "서버 Crash 감지"
        if cls == "VALIDATION_GAP":
            return "엔드포인트 버그 (Validation 미수행)"
        if cls == "CONNECTION_ERROR":
            return "서버 미응답 (연결 끊김)"
        if cls == "GRACEFUL_REJECTION":
            return f"정상 거부{' — ' + msg[:60] if msg else ''}"
        if cls == "UNEXPECTED_SUCCESS":
            return "예상치 못한 성공 (Validation 누락 의심)"
        if outcome == "passed":
            return "정상 처리됨"
        return "알 수 없음"

    @staticmethod
    def _classify_probe(test: dict[str, Any]) -> str:
        if test.get("probe_classification"):
            cls = str(test["probe_classification"])
            # ConnectionError 후 서버 죽음 → CRASH_DETECTED 로 승격
            if cls == "CONNECTION_ERROR" and test.get("server_alive") is False:
                return "CRASH_DETECTED"
            return cls

        longrepr = str(test.get("longrepr") or "")
        blob = longrepr.lower()
        if "CRASH_DETECTED" in longrepr:
            return "CRASH_DETECTED"
        if "VALIDATION_GAP" in longrepr:
            return "VALIDATION_GAP"
        # ConnectionResetError 10054 = 원격 호스트가 연결을 강제로 끊음 → 서버 Crash 의심
        if "connectionreset" in blob or "10054" in blob or "원격 호스트" in longrepr:
            if not test.get("server_alive", True):
                return "CRASH_DETECTED"
            return "CONNECTION_ERROR"

        outcome = str(test.get("outcome", "")).lower()
        if outcome == "passed":
            return "GRACEFUL_REJECTION"
        if outcome == "skipped":
            return "SKIPPED"
        return "OTHER_FAILURE"

    def _build_crash_probe_sheet(self, ws, probe_tests: list[dict[str, Any]]) -> None:
        self._title_banner(ws, "A1:M1", "Crash Probe — Robustness Layer (B)")
        ws.row_dimensions[1].height = 26

        crash_c = sum(1 for t in probe_tests if self._classify_probe(t) == "CRASH_DETECTED")
        gap_c = sum(1 for t in probe_tests if self._classify_probe(t) == "VALIDATION_GAP")
        graceful_c = sum(1 for t in probe_tests if self._classify_probe(t) == "GRACEFUL_REJECTION")
        other_c = len(probe_tests) - crash_c - gap_c - graceful_c

        kpi = [
            ("전체 Probe", len(probe_tests), "BDD7EE", "1F497D"),
            ("CRASH_DETECTED", crash_c, "F4CCCC", "C00000"),
            ("VALIDATION_GAP", gap_c, "FCE5CD", "7F3F00"),
            ("GRACEFUL_REJECTION", graceful_c, "C6EFCE", "375623"),
            ("기타", other_c, "EFEFEF", "404040"),
        ]

        for ci, (label, value, bg, fc) in enumerate(kpi, start=1):
            lc = ws.cell(row=2, column=ci, value=label)
            lc.font = Font(bold=True, color=_WHITE, size=9)
            lc.fill = PatternFill("solid", start_color=_BLUE_DARK, end_color=_BLUE_DARK)
            lc.alignment = Alignment(horizontal="center", vertical="center")
            lc.border = _BORDER
            vc = ws.cell(row=3, column=ci, value=value)
            vc.font = Font(bold=True, size=14, color=fc)
            vc.fill = PatternFill("solid", start_color=bg, end_color=bg)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border = _BORDER

        ws.row_dimensions[4].height = 6

        headers = [
            "#", "Endpoint", "Target Field", "Probe Label", "Probe Input",
            "Severity", "HTTP", "success", "error_code", "msg",
            "Classification", "Outcome", "Failure Reason"
        ]
        self._header_row(ws, 5, headers)

        for i, w in enumerate([5, 24, 16, 22, 42, 10, 8, 10, 12, 30, 18, 10, 40], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        if not probe_tests:
            ws.cell(row=6, column=1, value="(Crash Probe 결과 없음 — 파일 미존재 또는 job 미실행)")
            return

        for idx, t in enumerate(probe_tests, start=1):
            cls = t.get("probe_classification") or self._classify_probe(t)
            bg = self._PROBE_CLS_COLORS.get(cls, "EFEFEF")

            row_vals = [
                idx,
                t.get("probe_endpoint", ""),
                t.get("target_param", ""),
                t.get("probe_label", ""),
                self._format_value(t.get("probe_input")),
                t.get("probe_severity", ""),
                t.get("actual_status", ""),
                ("true" if t.get("response_success") is True else "false" if t.get("response_success") is False else "-"),
                t.get("response_error_code", ""),
                t.get("response_msg", ""),
                cls,
                str(t.get("outcome", "")).upper(),
                self._probe_failure_reason(t, cls),
            ]

            r = idx + 5
            for ci, val in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=ci, value=val)
                c.border = _BORDER
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.fill = PatternFill("solid", start_color=bg, end_color=bg)

        ws.freeze_panes = "A6"
        ws.auto_filter.ref = f"A5:M{len(probe_tests) + 5}"

    @staticmethod
    def _format_value(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value
        try:
            return json.dumps(value, ensure_ascii=False, indent=2)
        except Exception:
            return repr(value)

    def _build_actual_resp(self, item: dict[str, Any]) -> str:
        parts = []
        rs = item.get("response_success")
        if rs is not None:
            parts.append(f"success={'true' if rs else 'false'}")
        ec = item.get("response_error_code")
        if ec not in (None, ""):
            parts.append(f"error_code={ec}")
        dec = item.get("response_data_error_code")
        if dec not in (None, ""):
            parts.append(f"data.error_code={dec}")
        ms = item.get("response_data_match_score")
        if ms not in (None, ""):
            parts.append(f"data.match_score={ms}")
        ds = item.get("response_data_status")
        if ds not in (None, ""):
            parts.append(f"data.status={ds}")
        msg = item.get("response_msg") or ""
        if msg:
            parts.append(str(msg)[:40])
        return " / ".join(parts)

    def _title_banner(self, ws, cell_range: str, text: str) -> None:
        ws.merge_cells(cell_range)
        c = ws[cell_range.split(":")[0]]
        c.value = text
        c.font = Font(bold=True, size=13, color=_WHITE)
        c.fill = PatternFill("solid", start_color=_BLUE_TITLE, end_color=_BLUE_TITLE)
        c.alignment = Alignment(horizontal="center", vertical="center")

    def _section_header(self, ws, cell_range: str, text: str) -> None:
        ws.merge_cells(cell_range)
        c = ws[cell_range.split(":")[0]]
        c.value = text
        c.font = Font(bold=True, size=11, color=_BLUE_DARK)
        c.fill = PatternFill("solid", start_color=_BLUE_LIGHT, end_color=_BLUE_LIGHT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _BORDER

    def _header_row(self, ws, row: int, headers: list[str]) -> None:
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = Font(bold=True, color=_WHITE, size=9)
            c.fill = PatternFill("solid", start_color=_BLUE_DARK, end_color=_BLUE_DARK)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _BORDER
