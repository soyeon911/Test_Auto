from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
try:
    from tests.helpers.diag import classify_result as _classify_result
except ImportError:
    _classify_result = None

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─── colour palette ───────────────────────────────────────────────────────────
_BLUE_DARK = "1F497D"
_BLUE_TITLE = "2E75B6"
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_GREEN_BG = "F1F8F1"   # very light green
_RED_BG = "F9EEEE"     # very light red/pink
_YELLOW_BG = "FFF7D6"  # soft pale yellow

# Failure Cause 색상
_FC_COLORS: dict[str, str] = {
    "PASS":                              "EAF7EA",
    "서버 Crash (5xx)":                  "F4CCCC",
    "서버 미응답":                        "F4CCCC",
    "상태 미충족 — USER_NOT_FOUND":         "FCE5CD",
    "엔드포인트 버그 (Validation 미수행)": "FFE5E5",
    "엔드포인트 버그 (도메인 검증 미수행)": "FFE5E5",
    "예상된 실패":                        "EAF7EA",
    "Probe Only":                        "FFF2CC",
    "TC 실패 (단언 오류)":               "FFF2CC",
    "알 수 없음":                         "EFEFEF",
    "TC 생성 오류 (문법/들여쓰기)": "F4CCCC",
    "예상치 못한 실패": "FCE5CD",
}
_FC_COLORS.update({
    "입력 누락 (Required Missing)": "FCE5CD",
    "타입 오류 (Type Mismatch)": "FCE5CD",
    "범위 오류 (Out of Range)": "FFF2CC",
    "포맷 오류 (Encoding/Decode)": "FFF2CC",
    "리소스 없음 (Not Found)": "FCE5CD",
    "중복 요청 (Duplicate)": "FCE5CD",
    "도메인 조건 실패 (Threshold/Match)": "FFE5E5",
    "정상적인 실패 응답": "EAF7EA",
    "PASS (상태 미충족)": "EAF7EA",
    "도메인 조건 실패 (No Face/Detect Fail)": "FFE5E5",
    "입력 파라미터 오류 (Invalid Parameter)": "FCE5CD",
})
# http_status mode failure level colors
_FC_COLORS.update({
    "HTTP_STATUS_MISMATCH":  "F4CCCC",
    "ERROR_CODE_MISMATCH":   "FCE5CD",
    "BODY_SCHEMA_MISMATCH":  "FFF2CC",
    "SERVER_CRASH":          "F4CCCC",
    "CONNECTION_REFUSED":    "F4CCCC",
    "PASS_WITH_LEGACY_HTTP": "FFF2CC",
    "UNKNOWN":               "EFEFEF",
})

def _map_qfe_error_code(error_code: Any, msg: str, path: str = "") -> str | None:
    try:
        ec = int(error_code)
    except Exception:
        return None

    msg_l = (msg or "").lower()
    path = path or ""

    # 1) state — DB/리소스/설정 없음
    _STATE = {
        -20: "DATABASE_NOT_EXIST",
        -21: "FAILED_FILE_ALREADY_EXIST",
        -22: "FAILED_CREATE_DATABASE_FILE",
        -23: "FAILED_CREATE_TABLE",
        -24: "FAILED_SET_OPTIONS",
        -25: "FAILED_INSERT_USER",
        -26: "FAILED_DELETE_USER",
        -27: "FAILED_CLEAR_DATABASE",
        -28: "USER_NOT_FOUND",
        -29: "DATABASE_NOT_LOADED",
        -30: "CONFIG_NOT_EXIST",
        -31: "FAILED_CONFIG_CREATE",
        -43: "TEMPLATE_NOT_FOUND",
        -91: "FILE_NOT_FOUND",
    }
    if ec in _STATE:
        return f"상태 미충족 — {_STATE[ec]}"

    # 2) domain — 알고리즘/비즈니스 로직 실패
    _DOMAIN = {
        -32: "SYS_PARAM_NOT_SUPPORT",
        -33: "SYS_PARAM_OUT_OF_RANGE",
        -34: "INVALID_USER_ID",
        -35: "INVALID_SUB_ID",
        -40: "MAX_TEMPLATE_LIMIT",
        -41: "ADD_TEMPLATE",
        -42: "UPDATE_TEMPLATE",
        -50: "ENROLL_DIFFERENT_FACE",
        -200: "FAILED_FACE_DETECT",
        -201: "INVALID_ROI_COORDINATE",
        -202: "ROI_OUT_OF_IMAGE_BOUNDARY",
        -300: "FAILED_CHECK_REAL_FACE",
        -400: "FAILED_ESTIMATE_HEAD_POSE",
        -401: "FAILED_OUT_OF_RANGE",
        -500: "FAILED_MASK_SEGMENTATION",
        -600: "FAILED_ESTIMATE_FACE_ATTR",
        -700: "FAILED_EXTRACT_TEMPLATE",
        -800: "FAILED_GET_THRESHOLD",
    }
    if ec in _DOMAIN:
        return f"도메인 실패 — {_DOMAIN[ec]}"

    # 3) request/schema
    if ec == -90:
        return "요청 오류 — INVALID_PARAMETER"

    # 4) system — SDK/라이선스/모델
    _SYSTEM = {
        -2: "MEM_ALLOC", -3: "CAPTURE_FRAME", -4: "INSTANCE_NOT_EXIST",
        -5: "LICENSE", -6: "BUSY", -7: "THREAD_INSUFFICIENT",
        -8: "INVALID_POINTER", -9: "SDK_INSTANCE_BUSY",
        -10: "CAMERA_NOT_EXIST", -11: "FAILED_SET_CAMERA", -12: "FAILED_UNSET_CAMERA",
        -60: "LICENSE_EXPIRED", -61: "LICENSE_SUSPENDED", -62: "LICENSE_INVALID_PRODUCT",
        -63: "LICENSE_ACCESS_NO_PERMISSION", -64: "LICENSE_TIME_SYNC",
        -65: "LICENSE_INVALID_KEY", -66: "LICENSE_INVALID_FILE",
        -67: "LICENSE_MACHINE_CHANGED", -68: "LICENSE_TIME_MODIFIED", -69: "LICENSE_METADATA",
        -100: "FAILED_LOAD_MODEL", -101: "FAILED_GPU_SET",
        -9999: "UNKNOWN",
    }
    if ec in _SYSTEM:
        return f"시스템 오류 — {_SYSTEM[ec]}"

    # 5) generic -1
    if ec == -1:
        if any(x in msg_l for x in ["required", "unmarshal", "json", "type", "invalid request"]):
            return "타입 오류 (Type Mismatch)"
        if any(x in msg_l for x in ["base64", "decode", "padding", "encoding"]):
            return "포맷 오류 (Encoding/Decode)"
        if "failed to detect face" in msg_l or "error code: -200" in msg_l:
            return "도메인 실패 — FAILED_FACE_DETECT"
        if path in ("/api/v2/verify-template", "/api/v2/verify") and "verification failed" in msg_l:
            return "상태 미충족 — USER_NOT_FOUND"
        if path == "/api/v2/delete" and "failed to delete user" in msg_l:
            return "상태 미충족 — USER_NOT_FOUND"
        if "failed to get user template" in msg_l or "template not found" in msg_l or "user not found" in msg_l:
            return "상태 미충족 — USER_NOT_FOUND"
        return "정상적인 실패 응답"

    return None


def classify_failure_cause_from_item(item: dict[str, Any]) -> str:
    outcome = str(item.get("outcome", "")).lower()
    axis = str(item.get("axis") or "")
    reason_code = str(item.get("reason_code") or "")
    error_detail = str(item.get("error_detail") or "")
    expected_result_type = str(item.get("expected_result_type") or "")
    path = str(item.get("request_path") or "")
    response_success = item.get("response_success")
    response_error_code = item.get("response_error_code")
    response_data_error_code = item.get("response_data_error_code")
    response_msg = str(item.get("response_msg") or "")
    msg = response_msg.lower()
    data_status = str(item.get("response_data_status") or "").lower()
    actual_status = str(item.get("actual_status") or "")

    exc_type = str(item.get("exception_type") or "")
    exc_msg = str(item.get("exception_message") or "")
    longrepr = str(item.get("longrepr") or "")
    blob = f"{exc_type} {exc_msg} {longrepr} {error_detail}".lower()

    # 0. PASS
    if outcome == "passed":
        if (
            response_success is False
            and (reason_code == "precondition_not_met" or error_detail.startswith("state."))
        ):
            return "PASS (상태 미충족)"
        return "PASS"

    # 1. 인프라 / 런타임
    if item.get("server_crashed") or "crash_detected" in blob or actual_status.startswith("5"):
        return "서버 Crash (5xx)"

    if "connection" in blob or "refused" in blob or "timeout" in blob:
        return "서버 미응답"

    # 2. 생성/문법 오류
    if "indentationerror" in blob or "syntaxerror" in blob:
        return "TC 생성 오류 (문법/들여쓰기)"

    # 3. Crash Probe 우선 판정
    if "validation_gap" in blob:
        return "엔드포인트 버그 (Validation 미수행)"
    
    # QFE error_code 기반 보정
    mapped = _map_qfe_error_code(response_error_code, response_msg)
    if mapped:
        return mapped

    mapped = _map_qfe_error_code(response_data_error_code, response_msg)
    if mapped:
        return mapped


    # 4. 상태 미충족 — msg/reason_code 기반 추가 분류
    if reason_code == "precondition_not_met" or error_detail.startswith("state."):
        return "상태 미충족 — USER_NOT_FOUND"
    if "failed to get user template" in msg or "template not found" in msg:
        return "상태 미충족 — TEMPLATE_NOT_FOUND"
    if "failed to delete user" in msg or "user not found" in msg:
        return "상태 미충족 — USER_NOT_FOUND"
    if "template extraction failed" in msg:
        return "상태 미충족 — TEMPLATE_NOT_FOUND"
    if (
        "verification failed" in msg
        and path in ("/api/v2/verify-template", "/api/v2/verify")
    ):
        return "상태 미충족 — USER_NOT_FOUND"

    # 5. response body 문구 기반 분류
    # 5-1. 필수값/누락
    if "missing" in msg or "required" in msg:
        return "입력 누락 (Required Missing)"

    # 5-2. 타입/직렬화 오류
    if (
        "cannot unmarshal" in msg
        or "type mismatch" in msg
        or "invalid type" in msg
        or "wrong type" in msg
    ):
        return "타입 오류 (Type Mismatch)"

    # 5-3. 포맷/인코딩 오류
    if (
        "base64" in msg
        or "decode" in msg
        or "padding" in msg
        or "encoding" in msg
        or "invalid request" in msg and "json" in msg
    ):
        return "포맷 오류 (Encoding/Decode)"

    # 5-4. 범위 오류
    if "range" in msg or "out of range" in msg or "must be between" in msg:
        return "범위 오류 (Out of Range)"

    # 5-5. 리소스 없음
    if "not found" in msg or "no such" in msg or "does not exist" in msg:
        return "리소스 없음 (Not Found)"

    # 5-6. 중복
    if "duplicate" in msg or "already exists" in msg:
        return "중복 요청 (Duplicate)"

    # 5-7. threshold / match / verify 도메인 실패
    if "threshold" in msg:
        return "도메인 조건 실패 (Threshold/Match)"
    

    # 7. expected_result_type + success 조합
    if expected_result_type == "probe_only":
        if "assert" in blob or "failed:" in blob:
            return "TC 실패 (단언 오류)"
        return "Probe Only"

    if expected_result_type == "expected_fail":
        if response_success is True:
            if axis == "schema":
                return "엔드포인트 버그 (Validation 미수행)"
            return "엔드포인트 버그 (도메인 검증 미수행)"
        if response_success is False:
            return "정상적인 실패 응답"

    if expected_result_type == "expected_pass":
        if response_success is False:
            return "예상치 못한 실패"

    # 8. data.status 기반 보정
    if data_status in {"fail", "success"}:
        if expected_result_type == "expected_fail" and data_status == "success":
            return "엔드포인트 버그 (도메인 검증 미수행)"
        if expected_result_type == "expected_pass" and data_status == "fail":
            return "도메인 조건 실패 (Threshold/Match)"

    # 9. assertion/pytest failure
    if "assert" in blob or "failed:" in blob:
        return "TC 실패 (단언 오류)"

    return "알 수 없음"

_AXIS_LABEL: dict[str, str] = {
    "schema":  "schema (구조 검증)",
    "domain":  "domain (값 의미 검증)",
    "state":   "state (상태 검증)",
    "runtime": "runtime (안정성 검증)",
}

_METHOD_COLORS = {
    "GET": "D9EAD3",
    "POST": "FCE5CD",
    "PUT": "FFF2CC",
    "DELETE": "F4CCCC",
    "PATCH": "EAD1DC",
}


class ExcelReportBuilder:
    def __init__(self, output_path: str | Path):
        self.output_path = Path(output_path)

    def build(
        self,
        runner_summary: dict[str, Any],
        pytest_json_path: str | Path | None = None,
        source_file: str = "",
        base_url: str = "",
        endpoints: list[dict] | None = None,
        allure_results_dir: str | Path | None = None,
        crash_probe_json_path: str | Path | None = None,
    ) -> Path:
        normalized_tests = self._load_test_results(
            pytest_json_path=pytest_json_path,
            allure_results_dir=allure_results_dir,
        )

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Summary"
        self._build_summary(ws1, runner_summary, source_file, base_url)

        ws2 = wb.create_sheet("API List")
        if endpoints is None:
            endpoints = self._load_endpoints_from_source(source_file)
        self._build_api_list(ws2, endpoints or [])

        ws3 = wb.create_sheet("TC Table")
        self._build_tc_table(ws3, normalized_tests, base_url)

        ws4 = wb.create_sheet("Crash Probe")
        self._build_crash_probe_sheet(ws4, crash_probe_json_path)

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_path)
        return self.output_path

    def _load_endpoints_from_source(self, source_file: str) -> list[dict]:
        """Load API endpoints for the API List sheet without importing APIParser."""
        if not source_file:
            return []

        p = Path(source_file)
        if not p.exists() or not p.is_file():
            print(f"[ExcelReporter] API List source not found: {source_file}")
            return []

        try:
            if p.suffix.lower() in {".yaml", ".yml"}:
                try:
                    import yaml
                except ImportError:
                    print("[ExcelReporter] PyYAML is required to read YAML swagger files")
                    return []
                spec = yaml.safe_load(p.read_text(encoding="utf-8"))
            else:
                spec = json.loads(p.read_text(encoding="utf-8"))
        except Exception as exc:
            print(f"[ExcelReporter] API List source read failed: {exc}")
            return []

        if not isinstance(spec, dict):
            return []

        definitions = spec.get("definitions", {}) or {}
        base_path = str(spec.get("basePath") or "").strip().rstrip("/")

        def resolve_ref(obj):
            if not isinstance(obj, dict):
                return obj
            ref = obj.get("$ref")
            if not ref or not isinstance(ref, str) or not ref.startswith("#/"):
                return obj

            node = spec
            for part in ref.lstrip("#/").split("/"):
                if not isinstance(node, dict):
                    return obj
                node = node.get(part, {})
            return node if isinstance(node, dict) else obj

        def resolve_schema(schema):
            if not isinstance(schema, dict):
                return schema

            schema = resolve_ref(schema)
            if not isinstance(schema, dict):
                return schema

            out = dict(schema)

            props = out.get("properties")
            if isinstance(props, dict):
                out["properties"] = {
                    name: resolve_schema(prop)
                    for name, prop in props.items()
                }

            items = out.get("items")
            if isinstance(items, dict):
                out["items"] = resolve_schema(items)

            return out

        def normalize_param(param):
            param = resolve_ref(param)
            if not isinstance(param, dict):
                return {}

            schema = param.get("schema")
            if not schema:
                schema = {
                    "type": param.get("type", "string")
                }
                for key in [
                    "format", "enum", "minimum", "maximum",
                    "minLength", "maxLength", "pattern", "items",
                ]:
                    if key in param:
                        schema[key] = param[key]

            return {
                "name": param.get("name", ""),
                "in": param.get("in", "query"),
                "required": param.get("required", False),
                "description": param.get("description", ""),
                "schema": resolve_schema(schema),
            }

        def parse_body_param(param):
            if not param:
                return None

            param = resolve_ref(param)
            schema = resolve_schema(param.get("schema", {}))
            if not schema:
                return None

            return {
                "content_type": "application/json",
                "required": param.get("required", False),
                "description": param.get("description", ""),
                "name": param.get("name", ""),
                "schema": schema,
            }

        endpoints: list[dict] = []
        http_methods = {"get", "post", "put", "patch", "delete", "head", "options"}

        for raw_path, path_item in (spec.get("paths") or {}).items():
            if not isinstance(path_item, dict):
                continue

            path = str(raw_path)
            if base_path and not path.startswith(base_path + "/"):
                path = base_path + path

            path_level_params = path_item.get("parameters", []) or []

            for method, operation in path_item.items():
                method_l = str(method).lower()
                if method_l not in http_methods or not isinstance(operation, dict):
                    continue

                raw_params = list(path_level_params) + list(operation.get("parameters", []) or [])

                non_body_params = []
                body_param = None
                for rp in raw_params:
                    rp = resolve_ref(rp)
                    if isinstance(rp, dict) and rp.get("in") == "body":
                        body_param = rp
                    else:
                        non_body_params.append(rp)

                responses = {}
                for code, resp in (operation.get("responses") or {}).items():
                    resp = resolve_ref(resp)
                    if not isinstance(resp, dict):
                        resp = {}
                    responses[str(code)] = {
                        "description": resp.get("description", ""),
                        "schema": resolve_schema(resp.get("schema", {})) if resp.get("schema") else None,
                    }

                endpoints.append({
                    "path": path,
                    "method": method_l,
                    "operation_id": operation.get("operationId", f"{method_l}_{path}"),
                    "summary": operation.get("summary", ""),
                    "description": operation.get("description", ""),
                    "tags": operation.get("tags", []),
                    "parameters": [normalize_param(p) for p in non_body_params],
                    "request_body": parse_body_param(body_param),
                    "responses": responses,
                })

        print(f"[ExcelReporter] API List loaded {len(endpoints)} endpoint(s) from {source_file}")
        return endpoints

    def _build_summary(self, ws, summary: dict[str, Any], source_file: str, base_url: str) -> None:
        self._title_banner(ws, "A1:B1", "AutoTC — Test Execution Summary")
        ws.row_dimensions[1].height = 28
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 55

        rows: list[tuple[str, Any]] = [
            ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Source File", source_file),
            ("Base URL", base_url),
            ("", ""),
            ("Total Tests", summary.get("total", 0)),
            ("Passed", summary.get("passed", 0)),
            ("Failed", summary.get("failed", 0)),
            ("Error", summary.get("error", 0)),
            ("Duration (s)", summary.get("duration_seconds", "")),
            ("Return Code", summary.get("return_code", summary.get("returncode", ""))),
        ]

        for i, (label, value) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=label).font = Font(bold=bool(label))
            cell = ws.cell(row=i, column=2, value=value)
            if label == "Passed" and isinstance(value, int) and value > 0:
                cell.fill = PatternFill(fill_type="solid", fgColor=_GREEN_BG)
                cell.font = Font(bold=True)
            elif label == "Failed" and isinstance(value, int) and value > 0:
                cell.fill = PatternFill(fill_type="solid", fgColor=_RED_BG)
                cell.font = Font(bold=True)

    def _build_api_list(self, ws, endpoints: list[dict]) -> None:
        self._title_banner(ws, "A1:I1", "API Endpoint List")
        ws.row_dimensions[1].height = 24

        headers = [
            "#", "Method", "Path", "Operation ID",
            "Description",
            "Required Params",
            "Optional Params",
            "Request Body Fields (* = required)",
            "Success Code(s)",
        ]
        self._header_row(ws, 2, headers)

        col_widths = [5, 10, 35, 30, 45, 30, 28, 38, 15]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for idx, ep in enumerate(endpoints, start=1):
            method = ep.get("method", "").upper()
            path = ep.get("path", "")
            op_id = ep.get("operation_id", "")
            desc = ep.get("description") or ep.get("summary") or ""

            params = ep.get("parameters", [])
            req_params = [
                f"{p['name']} ({(p.get('schema') or {}).get('type', '?')})"
                for p in params if p.get("required")
            ]
            opt_params = [
                f"{p['name']} ({(p.get('schema') or {}).get('type', '?')})"
                for p in params if not p.get("required")
            ]

            body_schema = self._get_request_body_schema(ep.get("request_body"))
            body_req_set = set(body_schema.get("required", []))
            body_fields = [
                f"{'*' if f in body_req_set else ' '} {f} ({(s or {}).get('type', '?')})"
                for f, s in (body_schema.get("properties") or {}).items()
            ]

            responses = ep.get("responses", {})
            success_codes = [str(k) for k in responses if str(k).startswith("2")]

            row_vals = [
                idx, method, path, op_id, desc,
                "\n".join(req_params) or "—",
                "\n".join(opt_params) or "—",
                "\n".join(body_fields) or "—",
                ", ".join(success_codes) or "200",
            ]

            r = idx + 2
            bg = _METHOD_COLORS.get(method, "FFFFFF")
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 2:
                    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
                    cell.font = Font(bold=True)

            line_count = max(len(req_params), len(opt_params), len(body_fields), 1)
            ws.row_dimensions[r].height = max(18, 15 * line_count)
            ws.freeze_panes = "A3"

    def _build_tc_table(self, ws, tests: list[dict[str, Any]], base_url: str) -> None:
        self._title_banner(ws, "A1:AL1", "TC (Test Case) Table")
        ws.row_dimensions[1].height = 24

        headers = [
            "#",
            "Test Environment",
            "Target Type",
            "HTTP Method / Function",
            "Path / Module",
            "Rule Type",
            "Rule Subtype",
            "Endpoint Profile",
            "Expected Result Type",
            "Semantic Tag",
            "Policy",
            "Axis",
            "Reason Code",
            "Target Field",
            "Test Condition",
            "Request Query",
            "Request Headers",
            "Request Body / Arguments",
            "Expected",
            "Actual",
            "Response Msg",
            "Data Error Code",
            "Match Score",
            "Match Status",
            "Exception Type",
            "Exception Message",
            "Server Crash",
            "Server Log Tail",
            "Outcome",
            "Failure Cause",
            "Duration (s)",
            "Error Detail",
            "Failure Detail (pytest)",
            # http_status mode columns
            "HTTP Match",
            "Expected HTTP",
            "Expected Error Codes",
            "Error Code Match",
            "Failure Level",
        ]
        self._header_row(ws, 2, headers)

        col_widths = [
            5, 26, 12, 22, 34,
            16, 18, 18, 18, 16, 12,
            22, 22, 18, 42,
            28, 28, 70,
            28, 44, 38, 16, 16, 16,
            18, 42, 12, 42,
            12, 30, 12, 42, 55,
            12, 10, 28, 14, 24,
        ]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        src_cache: dict[str, str] = {}

        for idx, item in enumerate(tests, start=1):
            nodeid = item.get("nodeid", "")
            outcome = str(item.get("outcome", "unknown")).lower()
            duration = round(float(item.get("duration", 0) or 0), 3)
            longrepr = str(item.get("longrepr") or "")

            info = self._parse_nodeid(nodeid, src_cache)

            _raw_status = item.get("actual_status")
            if _raw_status is not None and _raw_status != "":
                actual_status = str(_raw_status)
            else:
                actual_status = self._extract_actual_status(longrepr)

            if not actual_status and outcome == "passed":
                actual_status = self._first_status_code(
                    item.get("expected_status_display") or info.get("expected_status", "")
                )

            target_type = item.get("target_type") or self._infer_target_type(item, info)
            method_or_function = item.get("request_method") or item.get("function") or info["method"]
            path_or_module = item.get("request_path") or info["path"] or item.get("request_url") or ""

            request_query = self._format_value(item.get("request_query", {}))
            request_headers = self._format_value(item.get("request_headers", {}))
            request_body_or_args = self._pick_request_payload(item)

            expected_display = item.get("expected_status_display") or self._build_expected_display(item, info)
            actual_display = self._build_actual_display(item, actual_status)

            rule_type = item.get("rule_type") or info["rule_type"]
            rule_subtype = item.get("rule_subtype", "")
            endpoint_profile = item.get("endpoint_profile", "")
            expected_result_type = item.get("expected_result_type", "")
            semantic_tag = item.get("semantic_tag", "")
            policy = item.get("policy", "")

            axis = item.get("axis", "")
            axis_display = _AXIS_LABEL.get(axis, axis)

            target_field = item.get("target_param", "")
            condition = item.get("condition") or info["condition"]

            response_msg = item.get("response_msg", "") or ""
            failure_cause = classify_failure_cause_from_item(item)

            # -- classify_result for http_status columns ------------------
            _exp_http = item.get("expected_http")
            _exp_ec   = item.get("expected_error_codes") or []
            _act_http = item.get("actual_status")
            _act_ec   = item.get("response_error_code")
            try:
                _act_http_int = int(str(_act_http)) if _act_http not in (None, "") else None
            except (ValueError, TypeError):
                _act_http_int = None
            try:
                _exp_http_int = int(str(_exp_http)) if _exp_http not in (None, "") else None
            except (ValueError, TypeError):
                _exp_http_int = None
            _http_match = (
                "YES" if _exp_http_int is not None and _act_http_int is not None and _exp_http_int == _act_http_int
                else "NO" if _exp_http_int is not None and _act_http_int is not None
                else "N/A"
            )
            _ec_match = (
                "YES" if _exp_ec and _act_ec is not None and _act_ec in _exp_ec
                else "NO" if _exp_ec and _act_ec is not None
                else "N/A"
            )
            _failure_level = ""
            if _classify_result is not None and outcome != "unknown":
                try:
                    _cr = _classify_result(
                        outcome=outcome,
                        expected_http=_exp_http,
                        expected_error_codes=_exp_ec or None,
                        actual_status=_act_http_int,
                        actual_error_code=_act_ec,
                        axis=axis,
                        reason_code=item.get("reason_code", ""),
                        server_crash=bool(item.get("server_crashed")),
                    )
                    _failure_level = _cr["level"]
                    if _cr.get("migration_flag"):
                        _failure_level += f" ({_cr['migration_flag']})"  
                except Exception:
                    pass

            row_vals = [
                idx,
                base_url,
                target_type,
                method_or_function,
                path_or_module,
                rule_type,
                rule_subtype,
                endpoint_profile,
                expected_result_type,
                semantic_tag,
                policy,
                axis_display,
                item.get("reason_code", ""),
                target_field,
                condition,
                request_query,
                request_headers,
                request_body_or_args,
                expected_display,
                actual_display,
                response_msg,
                item.get("response_data_error_code", ""),
                item.get("response_data_match_score", ""),
                item.get("response_data_status", ""),
                item.get("exception_type", ""),
                item.get("exception_message", ""),
                "Y" if item.get("server_crashed") else "",
                (item.get("server_log_tail") or "")[:2000],
                outcome.upper(),
                failure_cause,
                duration,
                item.get("error_detail", ""),
                longrepr[:2000] if outcome in {"failed", "broken"} else "",
                # http_status mode columns
                _http_match,
                str(_exp_http) if _exp_http is not None else "",
                str(sorted(_exp_ec)) if _exp_ec else "",
                _ec_match,
                _failure_level,
            ]

            r = idx + 2
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = _BORDER

            bg = _FC_COLORS.get(failure_cause, _YELLOW_BG)
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = PatternFill(fill_type="solid", fgColor=bg)

            fc_cell = ws.cell(row=r, column=30)
            if failure_cause not in {"PASS", "알 수 없음"}:
                fc_cell.font = Font(bold=True)

            max_len = max(
                len(str(v)) if v else 0
                for v in [
                    request_query,
                    request_body_or_args,
                    response_msg,
                    item.get("exception_message", ""),
                    longrepr,
                ]
            )
            if max_len > 300:
                height = 120
            elif max_len > 150:
                height = 80
            elif max_len > 80:
                height = 50
            else:
                height = 38
            ws.row_dimensions[r].height = height
            ws.freeze_panes = "A3"

    def _load_test_results(self, pytest_json_path: str | Path | None, allure_results_dir: str | Path | None) -> list[dict[str, Any]]:
        """테스트 결과를 로드한다.

        우선순위:
          1. report.json (pytest-json-report) — run-pipeline 아티팩트에서 다운로드된 최신 파일
          2. allure-results/ — report.json 이 없거나 전체가 인프라 실패(서버 미구동)인 경우 fallback

        pytest_report.json 은 구 버전 파이프라인 잔재이므로 절대 사용하지 않는다.
        generate-report job 이 pytest-raw 아티팩트를 먼저 다운로드하므로
        report.json 은 항상 현재 실행 결과를 담고 있다.
        """
        tests: list[dict[str, Any]] = []
        if pytest_json_path:
            p = Path(pytest_json_path)
            if p.exists() and p.is_file():
                try:
                    raw = json.loads(p.read_text(encoding="utf-8"))
                    if self._is_pytest_json_report(raw):
                        tests = self._normalize_pytest_json_report(raw)
                except Exception:
                    pass

        # allure-results fallback:
        #   - report.json 이 없거나 테스트가 0건인 경우
        #   - 또는 모든 테스트가 서버 미구동(ConnectionError)으로 인한 인프라 실패인 경우
        #   (allure-results 는 --alluredir 옵션으로 매 실행 전 초기화 후 생성됨)
        if allure_results_dir and self._should_fallback_to_allure(tests):
            d = Path(allure_results_dir)
            if d.exists() and d.is_dir():
                allure_tests = self._normalize_allure_results_dir(d)
                if allure_tests:
                    print(f"[ExcelReporter] allure fallback: {len(tests)} pytest → {len(allure_tests)} allure")
                    tests = allure_tests

        diag_map = self._load_diag_jsonl(pytest_json_path)
        if diag_map:
            for t in tests:
                diag = diag_map.get(t.get("nodeid", ""))
                if diag:
                    self._apply_diag(t, diag)

        return tests

    @staticmethod
    def _should_fallback_to_allure(tests: list[dict[str, Any]]) -> bool:
        """report.json 결과가 없거나 전부 인프라 실패(서버 미구동)이면 True."""
        if not tests:
            return True
        infra_types = {"ConnectionError", "ConnectionRefusedError", "NewConnectionError", "MaxRetryError"}
        return all(
            (t.get("exception_type") or "") in infra_types
            or "connection_refused" in (t.get("reason_code") or t.get("error_detail") or "").lower()
            for t in tests
        )

    @staticmethod
    def _load_diag_jsonl(pytest_json_path: str | Path | None) -> dict[str, dict]:
        candidates: list[Path] = []
        if pytest_json_path:
            candidates.append(Path(pytest_json_path).parent / "test_diag.jsonl")
        candidates.append(Path("reports/test_diag.jsonl"))

        for p in candidates:
            if not p.exists():
                continue
            result: dict[str, dict] = {}
            try:
                for line in p.read_text(encoding="utf-8").splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    rec = json.loads(line)
                    test_id = rec.get("test_id", "")
                    diag = rec.get("diag")
                    if test_id and isinstance(diag, dict):
                        result[test_id] = diag
                return result
            except Exception:
                pass
        return {}

    @staticmethod
    def _apply_diag(item: dict[str, Any], diag: dict[str, Any]) -> None:
        item["axis"] = diag.get("axis", "")
        item["reason_code"] = diag.get("reason_code", "")
        item["target_param"] = diag.get("target_field", item.get("target_param", ""))
        item["condition"] = diag.get("test_condition", item.get("condition", ""))

        exp_http = diag.get("expected_http") or ""
        exp_app = diag.get("expected_app") or ""
        parts = [p for p in [exp_http, exp_app] if p]
        if parts:
            item["expected_status_display"] = " / ".join(parts)

        if diag.get("actual_status") is not None:
            item["actual_status"] = str(diag["actual_status"])
        if diag.get("expected_http") is not None:
            item["expected_http"] = diag["expected_http"]

        req_body = diag.get("request_body")
        if req_body is not None:
            item["request_body"] = req_body
        req_query = diag.get("request_query")
        if req_query is not None:
            item["request_query"] = req_query
        req_headers = diag.get("request_headers")
        if req_headers is not None:
            item["request_headers"] = req_headers

        item["response_text"] = diag.get("response_snippet", item.get("response_text", ""))
        item["exception_type"] = diag.get("exception_type", "")
        item["exception_message"] = diag.get("exception_message", "")
        item["server_crashed"] = bool(diag.get("server_crash", False))
        item["error_detail"] = diag.get("error_detail", "")
        item["response_success"] = diag.get("response_success")
        item["response_error_code"] = diag.get("response_error_code")
        item["response_msg"] = diag.get("response_msg")
        item["response_data"] = diag.get("response_data")
        item["response_data_error_code"] = diag.get("response_data_error_code")
        item["response_data_match_score"] = diag.get("response_data_match_score")
        item["response_data_status"] = diag.get("response_data_status")
        item["response_data_verified"] = diag.get("response_data_verified")
        item["expected_error_codes"] = diag.get("expected_error_codes")
        item["expected_error_family"] = diag.get("expected_error_family", "")
        item["probe_endpoint"] = diag.get("probe_endpoint", item.get("probe_endpoint", ""))
        item["probe_label"] = diag.get("probe_label", item.get("probe_label", ""))
        item["probe_input"] = diag.get("probe_input", item.get("probe_input"))
        item["probe_severity"] = diag.get("probe_severity", item.get("probe_severity", ""))
        item["probe_classification"] = diag.get("probe_classification", item.get("probe_classification", ""))

    @staticmethod
    def _is_pytest_json_report(raw):
        return isinstance(raw, dict) and "tests" in raw

    def _normalize_pytest_json_report(self, raw):
        out = []
        for t in raw.get("tests", []):
            call = t.get("call") or {}
            meta = self._extract_tc_meta(t)

            # diag 가 user_properties 에 직접 포함된 경우 즉시 적용
            inline_diag: dict[str, Any] = meta.pop("_diag_from_up", {}) or {}

            item: dict[str, Any] = {
                "nodeid": t.get("nodeid", ""),
                "outcome": t.get("outcome", "unknown"),
                "duration": call.get("duration", t.get("duration", 0)),
                "longrepr": str(call.get("longrepr") or t.get("longrepr") or ""),
                "target_type": self._detect_target_type_from_meta(meta),
                "rule_type": meta.get("rule_type", ""),
                "rule_subtype": meta.get("rule_subtype", ""),
                "endpoint_profile": meta.get("endpoint_profile", ""),
                "semantic_tag": meta.get("semantic_tag", ""),
                "policy": meta.get("policy", ""),
                "expected_result_type": meta.get("expected_result_type", ""),
                "target_param": meta.get("target_param", ""),
                "condition": meta.get("condition", ""),
                "request_method": meta.get("request_method", ""),
                "request_url": meta.get("request_url", ""),
                "request_path": meta.get("request_path", ""),
                "request_path_params": meta.get("request_path_params", {}),
                "request_query": meta.get("request_query", {}),
                "request_headers": meta.get("request_headers", {}),
                "request_body": meta.get("request_body"),
                "function": meta.get("function", ""),
                "arguments": meta.get("arguments"),
                "arguments_repr": meta.get("arguments_repr", ""),
                "expected_status": meta.get("expected_status", []),
                "expected_status_display": self._coerce_expected_display(meta),
                "actual_status": meta.get("actual_status", ""),
                "response_text": meta.get("response_text", ""),
                "actual_result_repr": meta.get("actual_result_repr", ""),
                "actual_outcome": meta.get("actual_outcome", ""),
                "exception_type": meta.get("exception_type", ""),
                "exception_message": meta.get("exception_message", ""),
                "server_crashed": meta.get("server_crashed", False),
                "server_log_tail": meta.get("server_log_tail", ""),
                "axis": "",
                "reason_code": "",
                "error_detail": "",
                "response_success": None,
                "response_error_code": None,
                "response_msg": None,
                "response_data": None,
                "response_data_error_code": None,
                "response_data_match_score": None,
                "response_data_status": None,
                "expected_error_codes": meta.get("expected_error_codes"),   # tc_meta 직접 연결
                "expected_error_family": meta.get("expected_error_family", ""),
                "expected_http": meta.get("expected_http"),                    # tc_meta 직접 연결
            }
            # user_properties 에 diag 가 있으면 test_diag.jsonl 없이도 필드 채움
            if inline_diag:
                self._apply_diag(item, inline_diag)
            out.append(item)
        return out

    def _normalize_allure_results_dir(self, d: Path) -> list[dict[str, Any]]:
        out: list[dict[str, Any]] = []
        for p in sorted(d.glob("*-result.json")):
            try:
                raw = json.loads(p.read_text(encoding="utf-8"))
            except Exception:
                continue

            full_name = raw.get("fullName") or raw.get("name") or p.stem
            status = str(raw.get("status", "unknown")).lower()

            start = raw.get("start")
            stop = raw.get("stop")
            duration = 0.0
            if isinstance(start, (int, float)) and isinstance(stop, (int, float)) and stop >= start:
                duration = round((stop - start) / 1000.0, 3)

            details = raw.get("statusDetails", {}) or {}
            longrepr = details.get("message") or details.get("trace") or ""

            out.append({
                "nodeid": full_name,
                "outcome": status,
                "duration": duration,
                "longrepr": str(longrepr),
                "target_type": "",
                "rule_type": "",
                "rule_subtype": "",
                "endpoint_profile": "",
                "semantic_tag": "",
                "policy": "",
                "expected_result_type": "",
                "target_param": "",
                "condition": "",
                "request_method": "",
                "request_url": "",
                "request_path": "",
                "request_path_params": {},
                "request_query": {},
                "request_headers": {},
                "request_body": None,
                "function": "",
                "arguments": None,
                "arguments_repr": "",
                "expected_status": [],
                "expected_status_display": "",
                "actual_status": "",
                "response_text": "",
                "actual_result_repr": "",
                "actual_outcome": "",
                "exception_type": "",
                "exception_message": "",
                "server_crashed": False,
                "server_log_tail": "",
                "axis": "",
                "reason_code": "",
                "error_detail": "",
                "response_success": None,
                "response_error_code": None,
                "response_msg": None,
                "response_data": None,
                "response_data_error_code": None,
                "response_data_match_score": None,
                "response_data_status": None,
                "expected_error_codes": None,
                "expected_error_family": "",
            })
        return out

    @staticmethod
    def _first_status_code(text: str) -> str:
        parts = str(text or "").strip().split()
        if not parts:
            return ""
        return parts[0] if parts[0].isdigit() else ""

    @staticmethod
    def _extract_tc_meta(test_obj: dict[str, Any]) -> dict[str, Any]:
        meta: dict[str, Any] = {}
        md = test_obj.get("metadata") or {}
        if isinstance(md, dict):
            tc_meta = md.get("tc_meta")
            if isinstance(tc_meta, dict):
                meta.update(tc_meta)

        for up in test_obj.get("user_properties", []) or []:
            if isinstance(up, dict):
                if "tc_meta" in up and isinstance(up["tc_meta"], dict):
                    meta.update(up["tc_meta"])
                if "diag" in up and isinstance(up["diag"], dict):
                    meta["_diag_from_up"] = up["diag"]

                if "probe_meta" in up and isinstance(up["probe_meta"], dict):
                    meta.update(up["probe_meta"])
                if "probe_diag" in up and isinstance(up["probe_diag"], dict):
                    meta["_probe_diag_from_up"] = up["probe_diag"]

            elif isinstance(up, (list, tuple)) and len(up) == 2:
                if up[0] == "tc_meta" and isinstance(up[1], dict):
                    meta.update(up[1])
                if up[0] == "diag" and isinstance(up[1], dict):
                    meta["_diag_from_up"] = up[1]

                if up[0] == "probe_meta" and isinstance(up[1], dict):
                    meta.update(up[1])
                if up[0] == "probe_diag" and isinstance(up[1], dict):
                    meta["_probe_diag_from_up"] = up[1]

        return meta

    @staticmethod
    def _detect_target_type_from_meta(meta: dict[str, Any]) -> str:
        if meta.get("function"):
            return "python"
        if meta.get("request_method") or meta.get("request_body") is not None:
            return "api"
        return ""

    @staticmethod
    def _coerce_expected_display(meta: dict[str, Any]) -> str:
        if meta.get("expected_status_display"):
            return str(meta["expected_status_display"])
        if meta.get("expected_status"):
            return ", ".join(map(str, meta["expected_status"]))
        if meta.get("expected_exception_types"):
            return "raises " + ", ".join(map(str, meta["expected_exception_types"]))
        if meta.get("expected"):
            return str(meta["expected"])
        return ""

    def _build_expected_display(self, item: dict[str, Any], info: dict[str, str]) -> str:
        if item.get("expected_status_display"):
            return str(item["expected_status_display"])
        if item.get("expected_status"):
            return ", ".join(map(str, item["expected_status"]))
        return info.get("expected_status", "")

    def _build_actual_display(self, item: dict[str, Any], actual_status: str) -> str:
        head = str(actual_status) if actual_status else ""
        tail_parts: list[str] = []

        rs = item.get("response_success")
        if rs is not None:
            tail_parts.append(f"success={'true' if rs else 'false'}")

        ec = item.get("response_error_code")
        if ec not in [None, ""]:
            tail_parts.append(f"error_code={ec}")

        dec = item.get("response_data_error_code")
        if dec not in [None, ""]:
            tail_parts.append(f"data.error_code={dec}")

        ms = item.get("response_data_match_score")
        if ms not in [None, ""]:
            tail_parts.append(f"data.match_score={ms}")

        ds = item.get("response_data_status")
        if ds not in [None, ""]:
            tail_parts.append(f"data.status={ds}")

        verified = item.get("response_data_verified")
        if verified not in [None, ""]:
            tail_parts.append(f"data.verified={verified}")

        msg = item.get("response_msg")
        if msg:
            tail_parts.append(f"msg={str(msg)[:60]}")

        tail = ", ".join(tail_parts)
        if head and tail:
            return f"{head} / {tail}"
        return head or tail

    def _pick_request_payload(self, item: dict[str, Any]) -> str:
        if item.get("request_body") is not None:
            return self._format_value(item.get("request_body"))
        if item.get("arguments") is not None:
            return self._format_value(item.get("arguments"))
        if item.get("arguments_repr"):
            return str(item.get("arguments_repr"))
        return ""

    @staticmethod
    def _infer_target_type(item: dict[str, Any], info: dict[str, str]) -> str:
        if item.get("function"):
            return "python"
        if item.get("request_method") or info.get("method"):
            return "api"
        return ""

    @staticmethod
    def _format_value(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value
        try:
            return json.dumps(value, ensure_ascii=False, indent=2)
        except Exception:
            return repr(value)

    @staticmethod
    def _get_primary_media_type(req_body: dict | None) -> str | None:
        if not req_body:
            return None
        if "schema" in req_body:
            return "__legacy__"
        content = req_body.get("content", {})
        if not content:
            return None
        if "application/json" in content:
            return "application/json"
        if "multipart/form-data" in content:
            return "multipart/form-data"
        if "application/x-www-form-urlencoded" in content:
            return "application/x-www-form-urlencoded"
        return next(iter(content.keys()), None)

    def _get_request_body_schema(self, req_body: dict | None) -> dict[str, Any]:
        if not req_body:
            return {}
        media_type = self._get_primary_media_type(req_body)
        if not media_type:
            return {}
        if media_type == "__legacy__":
            return req_body.get("schema", {}) or {}
        return (req_body.get("content", {}).get(media_type, {}) or {}).get("schema", {}) or {}

    def _title_banner(self, ws, cell_range: str, text: str) -> None:
        ws.merge_cells(cell_range)
        first_cell = ws[cell_range.split(":")[0]]
        first_cell.value = text
        first_cell.font = Font(bold=True, size=13, color="FFFFFF")
        first_cell.fill = PatternFill(fill_type="solid", fgColor=_BLUE_TITLE)
        first_cell.alignment = Alignment(horizontal="center", vertical="center")
        first_cell.border = _BORDER

    def _header_row(self, ws, row: int, headers: list[str]) -> None:
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor=_BLUE_DARK)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _BORDER
        ws.row_dimensions[row].height = 18

    def _parse_nodeid(self, nodeid: str, src_cache: dict) -> dict:
        result: dict[str, str] = {
            "method": "",
            "path": "",
            "rule_type": "",
            "condition": "",
            "test_data": "",
            "expected_status": "",
        }
        if not nodeid:
            return result

        file_part = ""
        func_name = nodeid

        if "::" in nodeid:
            file_part, func_name = nodeid.split("::", 1)

        method, path = self._read_file_header(file_part, src_cache) if file_part else ("", "")
        result["method"] = method
        result["path"] = path

        if not method:
            m = re.search(r"test_(get|post|put|delete|patch)__", func_name, re.IGNORECASE)
            if m:
                result["method"] = m.group(1).upper()

        fn = func_name.lower()
        if fn.endswith("_positive"):
            result["rule_type"] = "positive"
            result["condition"] = "Happy path — all required fields present"
            result["test_data"] = "Representative valid values for all params"
            result["expected_status"] = "2xx"
        elif m2 := re.search(r"_missing_body_(.+)$", fn):
            field = m2.group(1)
            result["rule_type"] = "missing_required"
            result["condition"] = f"Required body field omitted: {field}"
            result["test_data"] = f"Request body without field '{field}'"
            result["expected_status"] = "400 / 422"
        elif fn.endswith("_missing_body"):
            result["rule_type"] = "missing_required"
            result["condition"] = "Required request body omitted"
            result["test_data"] = "No request body"
            result["expected_status"] = "400 / 415 / 422"
        elif m2 := re.search(r"_missing_(.+)$", fn):
            param = m2.group(1)
            result["rule_type"] = "missing_required"
            result["condition"] = f"Required param omitted: {param}"
            result["test_data"] = f"Request without param '{param}'"
            result["expected_status"] = "400 / 422"
        elif m2 := re.search(r"_wrong_type_body_(.+)$", fn):
            field = m2.group(1)
            result["rule_type"] = "wrong_type"
            result["condition"] = f"Body field sent with wrong type: {field}"
            result["test_data"] = f"body.{field} = wrong type"
            result["expected_status"] = "400 / 422"
        elif m2 := re.search(r"_wrong_type_(.+)$", fn):
            param = m2.group(1)
            result["rule_type"] = "wrong_type"
            result["condition"] = f"Param sent with wrong type: {param}"
            result["test_data"] = f"{param} = wrong type"
            result["expected_status"] = "400 / 422"
        elif m2 := re.search(r"_boundary_body_(.+?)_(.+)$", fn):
            field, probe = m2.group(1), m2.group(2)
            result["rule_type"] = "boundary"
            result["condition"] = f"Body field boundary probe: {field} ({probe})"
            result["test_data"] = f"body.{field} = {probe}"
            result["expected_status"] = "< 500"
        elif m2 := re.search(r"_boundary_(.+?)_(.+)$", fn):
            param, probe = m2.group(1), m2.group(2)
            result["rule_type"] = "boundary"
            result["condition"] = f"Param boundary probe: {param} ({probe})"
            result["test_data"] = f"{param} = {probe}"
            result["expected_status"] = "< 500"
        elif m2 := re.search(r"_invalid_enum_body_(.+)$", fn):
            field = m2.group(1)
            result["rule_type"] = "invalid_enum"
            result["condition"] = f"Body field outside allowed enum: {field}"
            result["test_data"] = f"body.{field} = '__INVALID_ENUM_VALUE__'"
            result["expected_status"] = "400 / 422"
        elif m2 := re.search(r"_invalid_enum_(.+)$", fn):
            param = m2.group(1)
            result["rule_type"] = "invalid_enum"
            result["condition"] = f"Param value outside allowed enum: {param}"
            result["test_data"] = f"{param} = '__INVALID_ENUM_VALUE__'"
            result["expected_status"] = "400 / 422"
        elif m2 := re.search(r"_semantic_(.+?)_(.+)$", fn):
            field, probe = m2.group(1), m2.group(2)
            result["rule_type"] = "semantic_probe"
            result["condition"] = f"Semantic probe: {field} ({probe})"
            result["test_data"] = f"{field} = {probe}"
            result["expected_status"] = "< 500"

        return result

    def _read_file_header(self, file_path: str, cache: dict) -> tuple[str, str]:
        if not file_path:
            return "", ""
        if file_path not in cache:
            p = Path(file_path)
            cache[file_path] = p.read_text(encoding="utf-8") if p.exists() else ""
        src = cache[file_path]
        m = re.search(r"^#\s*(GET|POST|PUT|DELETE|PATCH)\s+(\S+)", src, re.MULTILINE | re.IGNORECASE)
        if m:
            return m.group(1).upper(), m.group(2).strip()
        return "", ""

    def _extract_actual_status(self, longrepr: str) -> str:
        patterns = [
            r"\bgot\s+(\d{3})\b",
            r"\bstatus(?:_code)?\D+(\d{3})\b",
            r"\bHTTP[/ ]1\.[01]\"\s+(\d{3})\b",
        ]
        for pat in patterns:
            m = re.search(pat, longrepr, re.IGNORECASE)
            if m:
                return m.group(1)
        return ""

    # ─── Crash Probe sheet ────────────────────────────────────────────────────

    _PROBE_CLS_COLORS: dict[str, str] = {
        "CRASH_DETECTED":    "F4CCCC",
        "VALIDATION_GAP":    "FCE5CD",
        "GRACEFUL_REJECTION": "EAF7EA",
        "SKIPPED":           "FFF2CC",
        "OTHER_FAILURE":     "EFEFEF",
    }

    @staticmethod
    def _probe_failure_reason(t: dict, cls: str) -> str:
        """Crash Probe 행의 Failure Reason 컬럼 — 분류 기반 의미있는 설명."""
        ec = t.get("response_error_code")
        msg = str(t.get("response_msg") or "")
        outcome = str(t.get("outcome", "")).lower()
        _STATE = {
            -20: "DATABASE_NOT_EXIST", -21: "FAILED_FILE_ALREADY_EXIST",
            -22: "FAILED_CREATE_DATABASE_FILE", -23: "FAILED_CREATE_TABLE",
            -24: "FAILED_SET_OPTIONS", -25: "FAILED_INSERT_USER",
            -26: "FAILED_DELETE_USER", -27: "FAILED_CLEAR_DATABASE",
            -28: "USER_NOT_FOUND", -29: "DATABASE_NOT_LOADED",
            -30: "CONFIG_NOT_EXIST", -31: "FAILED_CONFIG_CREATE",
            -43: "TEMPLATE_NOT_FOUND", -91: "FILE_NOT_FOUND",
        }
        _DOMAIN = {
            -32: "SYS_PARAM_NOT_SUPPORT", -33: "SYS_PARAM_OUT_OF_RANGE",
            -34: "INVALID_USER_ID", -35: "INVALID_SUB_ID",
            -40: "MAX_TEMPLATE_LIMIT", -41: "ADD_TEMPLATE", -42: "UPDATE_TEMPLATE",
            -50: "ENROLL_DIFFERENT_FACE",
            -200: "FAILED_FACE_DETECT", -201: "INVALID_ROI_COORDINATE",
            -202: "ROI_OUT_OF_IMAGE_BOUNDARY", -300: "FAILED_CHECK_REAL_FACE",
            -400: "FAILED_ESTIMATE_HEAD_POSE", -401: "FAILED_OUT_OF_RANGE",
            -500: "FAILED_MASK_SEGMENTATION", -600: "FAILED_ESTIMATE_FACE_ATTR",
            -700: "FAILED_EXTRACT_TEMPLATE", -800: "FAILED_GET_THRESHOLD",
        }
        try:
            ec_int = int(ec)
            if ec_int in _STATE:
                return f"상태 미충족 — {_STATE[ec_int]}"
            if ec_int in _DOMAIN:
                return f"도메인 실패 — {_DOMAIN[ec_int]}"
            if ec_int == -90:
                return "요청 오류 — INVALID_PARAMETER"
            if ec_int == -1:
                return "정상적인 실패 응답 (ec=-1)"
        except (TypeError, ValueError):
            pass
        if cls == "CRASH_DETECTED":
            return "서버 Crash 감지"
        if cls == "VALIDATION_GAP":
            return "엔드포인트 버그 (Validation 미수행)"
        if cls == "CONNECTION_ERROR":
            return "서버 미응답 (연결 끊김)"
        if cls == "GRACEFUL_REJECTION":
            return f"정상 거부{chr(32) + chr(45) + chr(45) + chr(32) + msg[:60] if msg else ''}"
        if cls == "UNEXPECTED_SUCCESS":
            return "예상치 못한 성공 (Validation 누락 의심)"
        if outcome == "passed":
            return "정상 처리됨"
        return "알 수 없음"

    @staticmethod
    def _classify_probe(test: dict[str, Any]) -> str:
        if test.get("probe_classification"):
            cls = str(test["probe_classification"])
            # ConnectionError 후 서버 죽음 → CRASH_DETECTED 로 승격
            if cls == "CONNECTION_ERROR" and test.get("server_alive") is False:
                return "CRASH_DETECTED"
            return cls

        longrepr = str(test.get("longrepr") or "")
        blob = longrepr.lower()
        if "CRASH_DETECTED" in longrepr:
            return "CRASH_DETECTED"
        if "VALIDATION_GAP" in longrepr:
            return "VALIDATION_GAP"
        # ConnectionResetError 10054 = 원격 호스트가 연결을 강제로 끊음 → 서버 Crash 의심
        if "connectionreset" in blob or "10054" in blob or "원격 호스트" in longrepr:
            if not test.get("server_alive", True):
                return "CRASH_DETECTED"
            return "CONNECTION_ERROR"

        outcome = str(test.get("outcome", "")).lower()
        if outcome == "passed":
            return "GRACEFUL_REJECTION"
        if outcome == "skipped":
            return "SKIPPED"
        return "OTHER_FAILURE"
    
    
    def _normalize_probe_tests(self, raw: dict[str, Any]) -> list[dict[str, Any]]:
        out = []
        for t in raw.get("tests", []):
            call = t.get("call") or {}
            meta = self._extract_tc_meta(t)
            probe_diag = meta.pop("_probe_diag_from_up", {}) or {}

            item = {
                "nodeid": t.get("nodeid", ""),
                "outcome": t.get("outcome", "unknown"),
                "duration": call.get("duration", t.get("duration", 0)),
                "longrepr": str(call.get("longrepr") or t.get("longrepr") or ""),
                "target_param": meta.get("target_field", ""),
                "probe_endpoint": meta.get("probe_endpoint", ""),
                "probe_label": meta.get("probe_label", ""),
                "probe_input": meta.get("probe_input"),
                "probe_severity": meta.get("probe_severity", ""),
                "expected_result_type": meta.get("expected_result_type", "probe_only"),
                "actual_status": "",
                "response_success": None,
                "response_error_code": None,
                "response_msg": None,
                "error_detail": "",
                "axis": "runtime",
                "reason_code": "probe_runtime",
                "probe_classification": meta.get("probe_classification", ""),
            }

            if probe_diag:
                self._apply_diag(item, probe_diag)

            out.append(item)
        return out

    def _build_crash_probe_sheet(self, ws, report_path: str | Path | None) -> None:
        """Robustness Layer B — Crash Probe 결과 시트."""
        self._title_banner(ws, "A1:M1", "Crash Probe — Robustness Layer (B)")
        ws.row_dimensions[1].height = 24

        tests: list[dict[str, Any]] = []
        if report_path:
            p = Path(report_path)
            if p.exists():
                try:
                    raw = json.loads(p.read_text(encoding="utf-8"))
                    tests = self._normalize_probe_tests(raw)
                except Exception:
                    pass

        # ── KPI 요약 (row 2-4) ──────────────────────────────────────────────
        crash_count    = sum(1 for t in tests if self._classify_probe(t) == "CRASH_DETECTED")
        gap_count      = sum(1 for t in tests if self._classify_probe(t) == "VALIDATION_GAP")
        graceful_count = sum(1 for t in tests if self._classify_probe(t) == "GRACEFUL_REJECTION")
        other_count    = len(tests) - crash_count - gap_count - graceful_count

        kpi_labels  = ["전체 Probe", "CRASH_DETECTED", "VALIDATION_GAP", "GRACEFUL_REJECTION", "기타"]
        kpi_values  = [len(tests), crash_count, gap_count, graceful_count, other_count]
        kpi_colors  = ["BDD7EE",    "F4CCCC",        "FCE5CD",       "EAF7EA",           "EFEFEF"]
        kpi_fonts   = ["1F497D",    "C00000",         "7F3F00",       "375623",           "404040"]

        for ci, (label, value, bg, fc) in enumerate(zip(kpi_labels, kpi_values, kpi_colors, kpi_fonts), start=1):
            lc = ws.cell(row=2, column=ci, value=label)
            lc.font = Font(bold=True, color="FFFFFF")
            lc.fill = PatternFill(fill_type="solid", fgColor=_BLUE_DARK)
            lc.alignment = Alignment(horizontal="center", vertical="center")
            vc = ws.cell(row=3, column=ci, value=value)
            vc.font = Font(bold=True, size=14, color=fc)
            vc.fill = PatternFill(fill_type="solid", fgColor=bg)
            vc.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[4].height = 6  # separator

        # ── 결과 테이블 ─────────────────────────────────────────────────────
        headers = [
            "#", "Endpoint", "Target Field", "Probe Label", "Probe Input",
            "Severity", "HTTP", "success", "error_code", "msg",
            "Classification", "Outcome", "Failure Reason"
        ]
        self._header_row(ws, 5, headers)

        col_widths = [5, 24, 16, 22, 42, 10, 8, 10, 12, 30, 18, 10, 45]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        if not tests:
            ws.cell(row=6, column=1, value="(Crash Probe 결과 없음 — 파일 미존재 또는 job 미실행)")
            return

        for idx, t in enumerate(tests, start=1):
            cls = t.get("probe_classification") or self._classify_probe(t)
            bg = self._PROBE_CLS_COLORS.get(cls, "EFEFEF")

            row_vals = [
                idx,
                t.get("probe_endpoint", ""),
                t.get("target_param", ""),
                t.get("probe_label", ""),
                self._format_value(t.get("probe_input")),
                t.get("probe_severity", ""),
                t.get("actual_status", ""),
                ("true" if t.get("response_success") is True else "false" if t.get("response_success") is False else "-"),
                t.get("response_error_code", ""),
                t.get("response_msg", ""),
                cls,
                str(t.get("outcome", "")).upper(),
                self._probe_failure_reason(t, cls),
            ]

            r = idx + 5
            for ci, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.fill = PatternFill(fill_type="solid", fgColor=bg)
                cell.border = _BORDER

        ws.freeze_panes = "A6"

